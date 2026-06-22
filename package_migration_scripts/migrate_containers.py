#!/usr/bin/env python3
"""
GitLab Container Registry to GitHub Container Registry (GHCR) Migration Tool

Migrates container images (with all tags) from GitLab Container Registry to
GitHub Container Registry (ghcr.io) using skopeo or docker CLI.

Also verifies image digests (SHA-256) pre/post migration to detect integrity
changes (similar to the known GH-GH package SHA issue).

Prerequisites:
    - skopeo installed (preferred) OR docker/podman CLI available
    - Credentials configured for both registries

Usage:
    # Dry-run: list images that would be migrated
    python migrate_containers.py --gitlab-group ranjiths-infomagnus-group \
        --gitlab-token glpat-xxx --github-org my-gh-org --github-token ghp_xxx --dry-run

    # Migrate all images
    python migrate_containers.py --gitlab-group ranjiths-infomagnus-group \
        --gitlab-token glpat-xxx --github-org my-gh-org --github-token ghp_xxx

    # Migrate + verify digests
    python migrate_containers.py --gitlab-group ranjiths-infomagnus-group \
        --gitlab-token glpat-xxx --github-org my-gh-org --github-token ghp_xxx --verify

    # Use inventory Excel file instead of API discovery
    python migrate_containers.py --inventory gitlab-registry-inventory.xlsx \
        --gitlab-token glpat-xxx --github-org my-gh-org --github-token ghp_xxx

Environment variables:
    GITLAB_URL    - GitLab instance URL (default: https://gitlab.com)
    GITLAB_TOKEN  - GitLab PAT (read_registry scope)
    GITHUB_TOKEN  - GitHub PAT (write:packages scope)
    GITHUB_ORG    - GitHub organization
"""

import argparse
import json
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import requests

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION = "1.0.0"
DEFAULT_GITLAB_URL = "https://gitlab.com"
DEFAULT_GITLAB_REGISTRY = "registry.gitlab.com"
DEFAULT_GHCR = "ghcr.io"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_logger = logging.getLogger("container_migration")


def setup_logging(debug=False, log_file=None):
    level = logging.DEBUG if debug else logging.INFO
    _logger.setLevel(level)
    fmt = logging.Formatter("[%(asctime)s] %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(level)
    console.setFormatter(fmt)
    _logger.addHandler(console)

    if log_file:
        fh = logging.FileHandler(log_file, mode="w", encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        _logger.addHandler(fh)


def log(msg):
    _logger.info(msg)


def debug_log(msg):
    _logger.debug(msg)


def error_log(msg):
    _logger.error(msg)


def _short_error(exc, limit=200):
    """Return a concise error string for logs/reports."""
    text = str(exc).strip().replace("\n", " ")
    return text[:limit]


# ---------------------------------------------------------------------------
# .env autoload
# ---------------------------------------------------------------------------

def _load_dotenv():
    try:
        from dotenv import load_dotenv, find_dotenv
        env_file = find_dotenv(usecwd=True)
        if env_file:
            load_dotenv(env_file, override=False)
    except ImportError:
        pass


_load_dotenv()


# ---------------------------------------------------------------------------
# Tool Detection
# ---------------------------------------------------------------------------

def find_copy_tool():
    """Detect available container copy tool: skopeo > crane > docker."""
    if shutil.which("skopeo"):
        return "skopeo"
    if shutil.which("crane"):
        return "crane"
    if shutil.which("docker"):
        return "docker"
    if shutil.which("podman"):
        return "podman"
    return None


# ---------------------------------------------------------------------------
# GitLab Container Registry Client
# ---------------------------------------------------------------------------

class GitLabRegistryClient:
    """Interact with GitLab Container Registry API."""

    def __init__(self, gitlab_url, token, timeout=60, retries=3):
        self.base_url = gitlab_url.rstrip("/") + "/api/v4"
        self.token = token
        self.timeout = timeout
        self.retries = retries
        self.session = requests.Session()
        self.session.headers.update({"PRIVATE-TOKEN": token})

    def _get(self, path, params=None):
        url = self.base_url + path
        for attempt in range(self.retries + 1):
            try:
                resp = self.session.get(url, params=params, timeout=self.timeout)
                if resp.status_code == 429:
                    time.sleep(int(resp.headers.get("Retry-After", 2 ** (attempt + 1))))
                    continue
                if resp.status_code >= 500 and attempt < self.retries:
                    time.sleep(2 ** attempt)
                    continue
                return resp
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
                if attempt >= self.retries:
                    raise
                time.sleep(2 ** attempt)
        return resp

    def _get_paginated(self, path, params=None):
        results = []
        page = 1
        base_params = dict(params or {})
        while True:
            base_params["per_page"] = 100
            base_params["page"] = page
            resp = self._get(path, params=base_params)
            if resp.status_code == 404:
                return results
            resp.raise_for_status()
            try:
                items = resp.json()
            except ValueError as exc:
                raise RuntimeError(
                    f"Invalid JSON from GitLab for {path} page {page}: {_short_error(exc)}"
                ) from exc
            if not items:
                break
            results.extend(items)
            next_page = resp.headers.get("X-Next-Page")
            if not next_page:
                break
            page = int(next_page)
        return results

    def get_group_projects(self, group_path):
        """Get all projects under a group."""
        encoded = quote(group_path, safe="")
        return self._get_paginated(
            f"/groups/{encoded}/projects",
            params={"include_subgroups": "true", "with_shared": "false"},
        )

    def get_repositories(self, project_id):
        """Get container registry repositories for a project."""
        return self._get_paginated(
            f"/projects/{project_id}/registry/repositories",
            params={"tags_count": "true"},
        )

    def get_repository_tags(self, project_id, repository_id):
        """Get all tags for a container repository."""
        return self._get_paginated(
            f"/projects/{project_id}/registry/repositories/{repository_id}/tags",
        )

    def get_tag_detail(self, project_id, repository_id, tag_name):
        """Get detailed info for a specific tag (includes digest)."""
        resp = self._get(
            f"/projects/{project_id}/registry/repositories/{repository_id}/tags/{quote(tag_name, safe='')}"
        )
        if resp.status_code == 200:
            try:
                return resp.json()
            except ValueError:
                return None
        return None


# ---------------------------------------------------------------------------
# Container Copy Operations
# ---------------------------------------------------------------------------

class ContainerCopier:
    """Copy container images between registries using skopeo/crane/docker."""

    def __init__(self, tool, gitlab_token, github_token, github_org,
                 gitlab_registry=DEFAULT_GITLAB_REGISTRY, ghcr=DEFAULT_GHCR):
        self.tool = tool
        self.gitlab_token = gitlab_token
        self.github_token = github_token
        self.github_org = github_org
        self.gitlab_registry = gitlab_registry
        self.ghcr = ghcr
        self._crane_logged_in = False

    def _run_with_retries(self, cmd, timeout=600, retries=2):
        """Run a subprocess command with retry for transient failures."""
        last_result = None
        for attempt in range(retries + 1):
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
                if result.returncode == 0:
                    return True, result.stdout, result.stderr
                last_result = result
                if attempt < retries:
                    wait = 2 ** attempt
                    debug_log(
                        f"Command failed (attempt {attempt + 1}/{retries + 1}), retrying in {wait}s: {' '.join(cmd[:3])}"
                    )
                    time.sleep(wait)
            except subprocess.TimeoutExpired:
                if attempt >= retries:
                    return False, "", f"Command timed out after {timeout}s"
                time.sleep(2 ** attempt)
            except FileNotFoundError:
                return False, "", f"{self.tool} not found in PATH"

        stderr = last_result.stderr if last_result else "command_failed"
        stdout = last_result.stdout if last_result else ""
        return False, stdout, stderr

    def _ensure_crane_login(self):
        """Authenticate crane to both source and destination registries once."""
        if self.tool != "crane" or self._crane_logged_in:
            return True, ""

        login_gl = subprocess.run(
            ["crane", "auth", "login", self.gitlab_registry,
             "-u", "oauth2", "-p", self.gitlab_token],
            capture_output=True, text=True, timeout=30,
        )
        if login_gl.returncode != 0:
            return False, f"Crane GitLab login failed: {login_gl.stderr.strip()}"

        login_gh = subprocess.run(
            ["crane", "auth", "login", self.ghcr,
             "-u", self.github_org, "-p", self.github_token],
            capture_output=True, text=True, timeout=30,
        )
        if login_gh.returncode != 0:
            return False, f"Crane GHCR login failed: {login_gh.stderr.strip()}"

        self._crane_logged_in = True
        return True, ""

    def copy_image(self, source_image, dest_image):
        """
        Copy a single image:tag from GitLab to GHCR.
        Returns (success: bool, stdout: str, stderr: str).
        """
        src = f"docker://{source_image}"
        dst = f"docker://{dest_image}"

        if self.tool == "skopeo":
            cmd = [
                "skopeo", "copy",
                "--src-creds", f"oauth2:{self.gitlab_token}",
                "--dest-creds", f"{self.github_org}:{self.github_token}",
                src, dst,
            ]
        elif self.tool == "crane":
            ok, err = self._ensure_crane_login()
            if not ok:
                return False, "", err
            # crane requires separate login; use copy command
            cmd = ["crane", "copy", source_image, dest_image]
        elif self.tool in ("docker", "podman"):
            # docker requires pull + tag + push workflow
            return self._docker_copy(source_image, dest_image)
        else:
            return False, "", f"Unsupported tool: {self.tool}"

        return self._run_with_retries(cmd, timeout=600, retries=2)

    def _docker_copy(self, source_image, dest_image):
        """Docker/podman pull-tag-push workflow."""
        tool = self.tool

        # Login to GitLab registry
        login_gl = subprocess.run(
            [tool, "login", self.gitlab_registry,
             "-u", "oauth2", "-p", self.gitlab_token],
            capture_output=True, text=True, timeout=30,
        )
        if login_gl.returncode != 0:
            return False, "", f"GitLab login failed: {login_gl.stderr}"

        # Login to GHCR
        login_gh = subprocess.run(
            [tool, "login", self.ghcr,
             "-u", self.github_org, "-p", self.github_token],
            capture_output=True, text=True, timeout=30,
        )
        if login_gh.returncode != 0:
            return False, "", f"GHCR login failed: {login_gh.stderr}"

        # Pull
        pull = subprocess.run(
            [tool, "pull", source_image],
            capture_output=True, text=True, timeout=600,
        )
        if pull.returncode != 0:
            return False, "", f"Pull failed: {pull.stderr}"

        # Tag
        tag = subprocess.run(
            [tool, "tag", source_image, dest_image],
            capture_output=True, text=True, timeout=30,
        )
        if tag.returncode != 0:
            return False, "", f"Tag failed: {tag.stderr}"

        # Push
        push = subprocess.run(
            [tool, "push", dest_image],
            capture_output=True, text=True, timeout=600,
        )
        if push.returncode != 0:
            return False, "", f"Push failed: {push.stderr}"

        return True, push.stdout, ""

    def inspect_digest(self, image_ref):
        """Get the manifest digest of an image (for SHA verification)."""
        if self.tool == "skopeo":
            # Determine creds based on registry
            if self.ghcr in image_ref:
                creds = f"{self.github_org}:{self.github_token}"
            else:
                creds = f"oauth2:{self.gitlab_token}"

            cmd = [
                "skopeo", "inspect", "--raw",
                "--creds", creds,
                f"docker://{image_ref}",
            ]
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                if result.returncode == 0:
                    # Compute digest of the manifest
                    import hashlib
                    digest = hashlib.sha256(result.stdout.encode()).hexdigest()
                    return f"sha256:{digest}"
            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

            # Try with --format for digest
            cmd = [
                "skopeo", "inspect",
                "--creds", creds,
                f"docker://{image_ref}",
            ]
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                if result.returncode == 0:
                    info = json.loads(result.stdout)
                    return info.get("Digest", "")
            except (subprocess.TimeoutExpired, FileNotFoundError, json.JSONDecodeError):
                pass

        elif self.tool == "crane":
            ok, _ = self._ensure_crane_login()
            if not ok:
                return None
            cmd = ["crane", "digest", image_ref]
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                if result.returncode == 0:
                    return result.stdout.strip()
            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

        return None


# ---------------------------------------------------------------------------
# Migration Engine
# ---------------------------------------------------------------------------

class ContainerMigrationEngine:
    """Orchestrates container image migration from GitLab to GHCR."""

    def __init__(self, gitlab_client, copier, github_org, dry_run=False, verify=False,
                 strict_tag_check=False):
        self.gitlab = gitlab_client
        self.copier = copier
        self.github_org = github_org
        self.dry_run = dry_run
        self.verify = verify
        self.strict_tag_check = strict_tag_check
        self._source_tags = {}  # {dest_image_base: set(tags)}

        self.report = {
            "started_at": datetime.utcnow().isoformat(),
            "tool": copier.tool if copier else "none",
            "mode": "dry-run" if dry_run else "live",
            "images_found": 0,
            "tags_found": 0,
            "tags_migrated": 0,
            "tags_failed": 0,
            "tags_skipped": 0,
            "digest_matches": 0,
            "digest_mismatches": 0,
            "mismatches": [],
            "errors": [],
            "details": [],
            "tag_verification": {
                "enabled": strict_tag_check,
                "checked": 0,
                "missing_tags": [],
            },
        }

    def run_from_group(self, group_path):
        """Discover and migrate all container images in a GitLab group."""
        log(f"Discovering container images in group: {group_path}")

        try:
            projects = self.gitlab.get_group_projects(group_path)
        except Exception as exc:
            error_log(f"Failed to list projects for group {group_path}: {_short_error(exc)}")
            self.report["errors"].append({
                "source": group_path,
                "error": f"group_list_failed: {_short_error(exc, 500)}",
            })
            self._finalize()
            return
        log(f"Found {len(projects)} projects")

        for project in projects:
            self._process_project(project)

        self._finalize()

    def run_from_inventory(self, inventory_path):
        """Use an inventory Excel file to drive migration."""
        try:
            import openpyxl
        except ImportError:
            error_log("openpyxl required for inventory mode: pip install openpyxl")
            sys.exit(1)

        log(f"Loading inventory from: {inventory_path}")
        try:
            wb = openpyxl.load_workbook(inventory_path, read_only=True)
        except Exception as exc:
            error_log(f"Failed to load inventory file: {_short_error(exc)}")
            return

        # Look for Container Registry sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "container" in name.lower():
                sheet_name = name
                break

        if not sheet_name:
            error_log("No Container Registry sheet found in inventory")
            wb.close()
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            log("No container images in inventory")
            return

        headers = rows[0]
        path_idx = headers.index("project_path") if "project_path" in headers else None
        proj_id_idx = headers.index("project_id") if "project_id" in headers else None

        if path_idx is None or proj_id_idx is None:
            error_log("Could not find project_path/project_id columns in inventory")
            return

        # Get unique projects
        seen_projects = set()
        for row in rows[1:]:
            project_id = row[proj_id_idx]
            project_path = row[path_idx]
            if project_id not in seen_projects:
                seen_projects.add(project_id)
                # Fetch project details
                project = {"id": project_id, "path_with_namespace": project_path,
                           "path": project_path.split("/")[-1]}
                self._process_project(project)

        self._finalize()

    def _process_project(self, project):
        """Process all container repositories in a project."""
        project_id = project["id"]
        project_path = project.get("path_with_namespace", "")
        project_name = project.get("path", project_path.split("/")[-1])

        try:
            repositories = self.gitlab.get_repositories(project_id)
        except Exception as e:
            log(f"  SKIPPED {project_path}: cannot access container registry ({e})")
            self.report.setdefault("skipped_projects", []).append({"project": project_path, "reason": str(e)})
            return
        if not repositories:
            debug_log(f"  No container repositories in {project_path}")
            return

        log(f"\nProject: {project_path} ({len(repositories)} repositories)")
        self.report["images_found"] += len(repositories)

        for repo in repositories:
            repo_id = repo["id"]
            repo_path = repo.get("path", "")
            repo_location = repo.get("location", "")
            tag_count = repo.get("tags_count", 0)

            log(f"  Repository: {repo_path} ({tag_count} tags)")

            try:
                tags = self.gitlab.get_repository_tags(project_id, repo_id)
            except Exception as exc:
                error_log(f"  FAILED listing tags for {repo_path}: {_short_error(exc)}")
                self.report["errors"].append({
                    "source": repo_location,
                    "error": f"tag_list_failed: {_short_error(exc, 500)}",
                })
                continue
            if not tags:
                continue

            self.report["tags_found"] += len(tags)

            for tag in tags:
                tag_name = tag.get("name", "")
                self._migrate_tag(
                    project_id, project_path, project_name,
                    repo_id, repo_path, repo_location, tag_name
                )

    def _migrate_tag(self, project_id, project_path, project_name,
                     repo_id, repo_path, repo_location, tag_name):
        """Migrate a single image:tag."""
        # Source: registry.gitlab.com/group/project:tag
        source_image = f"{repo_location}:{tag_name}"

        # Track source tags for strict verification
        dest_repo_name = project_path.replace("/", "-").lower()
        dest_base = f"{DEFAULT_GHCR}/{self.github_org}/{dest_repo_name}"
        self._source_tags.setdefault(dest_base, set()).add(tag_name)

        # Destination: ghcr.io/github-org/project-name:tag
        # Flatten the path for GHCR (replace / with -)
        dest_repo_name = project_path.replace("/", "-").lower()
        dest_image = f"{DEFAULT_GHCR}/{self.github_org}/{dest_repo_name}:{tag_name}"

        # Get source digest from GitLab API
        try:
            tag_detail = self.gitlab.get_tag_detail(project_id, repo_id, tag_name)
            source_digest = tag_detail.get("digest", "") if tag_detail else ""
        except Exception as exc:
            debug_log(f"    Could not fetch source digest for {source_image}: {_short_error(exc)}")
            source_digest = ""

        detail = {
            "source": source_image,
            "destination": dest_image,
            "tag": tag_name,
            "source_digest": source_digest,
            "status": "pending",
        }

        if self.dry_run:
            log(f"    [DRY-RUN] {source_image} -> {dest_image}")
            detail["status"] = "dry-run"
            self.report["tags_skipped"] += 1
            self.report["details"].append(detail)
            return

        # Skip migration when destination tag already exists.
        existing_dest_digest = self.copier.inspect_digest(dest_image)
        if existing_dest_digest:
            log(f"    SKIPPED (already exists): {dest_image}")
            detail["status"] = "already-exists"
            detail["dest_digest"] = existing_dest_digest
            self.report["tags_skipped"] += 1

            if self.verify and source_digest:
                if existing_dest_digest == source_digest:
                    log(f"    Existing digest MATCH: {source_digest[:24]}...")
                    self.report["digest_matches"] += 1
                else:
                    log("    Existing digest MISMATCH!")
                    log(f"      Source: {source_digest}")
                    log(f"      Dest:   {existing_dest_digest}")
                    self.report["digest_mismatches"] += 1
                    self.report["mismatches"].append({
                        "source_image": source_image,
                        "dest_image": dest_image,
                        "source_digest": source_digest,
                        "dest_digest": existing_dest_digest,
                    })

            self.report["details"].append(detail)
            return

        # Copy image
        log(f"    Copying: {source_image} -> {dest_image}")
        success, stdout, stderr = self.copier.copy_image(source_image, dest_image)

        if success:
            log(f"    OK: {tag_name}")
            detail["status"] = "migrated"
            self.report["tags_migrated"] += 1

            # Verify digest if requested
            if self.verify and source_digest:
                dest_digest = self.copier.inspect_digest(dest_image)
                detail["dest_digest"] = dest_digest

                if dest_digest and dest_digest == source_digest:
                    log(f"    Digest MATCH: {source_digest[:24]}...")
                    self.report["digest_matches"] += 1
                elif dest_digest:
                    log(f"    Digest MISMATCH!")
                    log(f"      Source: {source_digest}")
                    log(f"      Dest:   {dest_digest}")
                    self.report["digest_mismatches"] += 1
                    self.report["mismatches"].append({
                        "source_image": source_image,
                        "dest_image": dest_image,
                        "source_digest": source_digest,
                        "dest_digest": dest_digest,
                    })
                else:
                    debug_log(f"    Could not verify dest digest for {dest_image}")
        else:
            error_log(f"    FAILED: {tag_name} - {stderr[:200]}")
            detail["status"] = "failed"
            detail["error"] = stderr[:500]
            self.report["tags_failed"] += 1
            self.report["errors"].append({
                "source": source_image,
                "dest": dest_image,
                "error": stderr[:500],
            })

        self.report["details"].append(detail)

    def _verify_target_tags(self):
        """Verify all source tags exist in destination registry."""
        missing_tags = []
        checked = 0

        for dest_base, expected_tags in self._source_tags.items():
            for tag in sorted(expected_tags):
                checked += 1
                dest_ref = f"{dest_base}:{tag}"
                digest = self.copier.inspect_digest(dest_ref) if self.copier else None
                if not digest:
                    missing_tags.append({
                        "dest_image": dest_ref,
                        "tag": tag,
                    })

        self.report["tag_verification"]["checked"] = checked
        self.report["tag_verification"]["missing_tags"] = missing_tags

        if missing_tags:
            log("\nSTRICT TAG CHECK FAILED: missing destination tags detected")
            for item in missing_tags:
                log(f"  Missing: {item['dest_image']}")
        else:
            log(f"\nStrict tag check PASSED: all {checked} tags verified in destination")

        return len(missing_tags) == 0

    def _finalize(self):
        """Save report and print summary."""
        strict_ok = True
        if self.strict_tag_check and not self.dry_run and self.copier:
            strict_ok = self._verify_target_tags()
        self.report["strict_tag_check_passed"] = strict_ok

        self.report["completed_at"] = datetime.utcnow().isoformat()

        # Save report
        report_path = Path("container_migration_report.json")
        tmp_fd, tmp_path = tempfile.mkstemp(prefix="container_report_", suffix=".json", dir=str(report_path.parent))
        try:
            with os.fdopen(tmp_fd, "w", encoding="utf-8") as f:
                json.dump(self.report, f, indent=2, default=str)
            os.replace(tmp_path, report_path)
            log(f"\nReport saved to: {report_path}")
        except Exception as exc:
            error_log(f"Failed to write report: {_short_error(exc)}")
            self.report["errors"].append({"source": "report", "error": f"report_write_failed: {_short_error(exc, 500)}"})
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass

        # Summary
        r = self.report
        log("\n" + "=" * 60)
        log("CONTAINER MIGRATION SUMMARY")
        log("=" * 60)
        log(f"  Tool used:           {r['tool']}")
        log(f"  Mode:                {r['mode']}")
        log(f"  Repositories found:  {r['images_found']}")
        log(f"  Tags found:          {r['tags_found']}")
        log(f"  Tags migrated:       {r['tags_migrated']}")
        log(f"  Tags failed:         {r['tags_failed']}")
        log(f"  Tags skipped:        {r['tags_skipped']}")

        if self.verify:
            log(f"  Digest matches:      {r['digest_matches']}")
            log(f"  Digest mismatches:   {r['digest_mismatches']}")

        if r["mismatches"]:
            log("\n  DIGEST MISMATCHES (integrity changed during migration):")
            for m in r["mismatches"]:
                log(f"    {m['source_image']}")
                log(f"      Source: {m['source_digest']}")
                log(f"      Dest:   {m['dest_digest']}")

        if r["errors"]:
            log(f"\n  ERRORS ({len(r['errors'])}):")
            for e in r["errors"][:10]:
                log(f"    {e['source']} -> {e.get('error', '')[:100]}")

        log("=" * 60)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Migrate container images from GitLab Container Registry to GHCR"
    )

    parser.add_argument("--gitlab-url", default=os.getenv("GITLAB_URL", DEFAULT_GITLAB_URL))
    parser.add_argument("--gitlab-token", default=os.getenv("GITLAB_TOKEN"),
                        help="GitLab PAT (read_registry scope)")
    parser.add_argument("--gitlab-group", default=None,
                        help="GitLab group path to discover images")
    parser.add_argument("--github-token", default=os.getenv("GITHUB_TOKEN"),
                        help="GitHub PAT (write:packages scope)")
    parser.add_argument("--github-org", default=os.getenv("GITHUB_ORG"),
                        help="GitHub organization (GHCR namespace)")

    parser.add_argument("--inventory", default=None,
                        help="Path to gitlab-registry-inventory.xlsx (alternative to --gitlab-group)")

    parser.add_argument("--dry-run", action="store_true",
                        help="List images without copying")
    parser.add_argument("--verify", action="store_true",
                        help="Verify digests after migration")
    parser.add_argument("--tool", choices=["skopeo", "crane", "docker", "podman"],
                        default=None, help="Force a specific copy tool")
    parser.add_argument("--use-docker", action="store_true",
                        help="Force Docker as the container copy tool")

    parser.add_argument("--strict-tag-check", action="store_true",
                        help="Fail run if any source tag is missing in destination registry")
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--log-file", default="container_migration.log")

    return parser.parse_args()


def main():
    try:
        args = parse_args()
        setup_logging(debug=args.debug, log_file=args.log_file)

        log(f"GitLab-to-GHCR Container Migration Tool v{VERSION}")

        # Validate
        if not args.gitlab_token:
            error_log("GitLab token required (--gitlab-token or GITLAB_TOKEN)")
            sys.exit(1)
        if not args.github_token and not args.dry_run:
            error_log("GitHub token required (--github-token or GITHUB_TOKEN)")
            sys.exit(1)
        if not args.github_org:
            error_log("GitHub org required (--github-org or GITHUB_ORG)")
            sys.exit(1)
        if not args.gitlab_group and not args.inventory:
            error_log("Specify either --gitlab-group or --inventory")
            sys.exit(1)

        # Detect copy tool
        tool = "docker" if args.use_docker else (args.tool or find_copy_tool())
        if not tool and not args.dry_run:
            error_log(
                "No container copy tool found. Install one of: skopeo, crane, docker, podman"
            )
            sys.exit(1)

        if tool:
            log(f"Using copy tool: {tool}")
        else:
            log("No copy tool available (dry-run mode, not needed)")

        # Initialize
        gitlab_client = GitLabRegistryClient(args.gitlab_url, args.gitlab_token)

        copier = None
        if tool:
            copier = ContainerCopier(
                tool=tool,
                gitlab_token=args.gitlab_token,
                github_token=args.github_token or "",
                github_org=args.github_org,
            )

        engine = ContainerMigrationEngine(
            gitlab_client=gitlab_client,
            copier=copier,
            github_org=args.github_org,
            dry_run=args.dry_run,
            verify=args.verify,
            strict_tag_check=args.strict_tag_check,
        )

        if args.inventory:
            engine.run_from_inventory(args.inventory)
        else:
            engine.run_from_group(args.gitlab_group)

        if args.strict_tag_check and not engine.report.get("strict_tag_check_passed", True):
            sys.exit(2)
    except KeyboardInterrupt:
        error_log("Interrupted by user")
        sys.exit(130)
    except Exception as exc:
        error_log(f"Fatal error: {_short_error(exc, 500)}")
        debug_log("Run with --debug for full trace details.")
        sys.exit(1)


if __name__ == "__main__":
    main()
