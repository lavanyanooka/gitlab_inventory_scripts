#!/usr/bin/env python3
"""
GitLab CI/CD Secrets and Variables Inventory Tool (Standalone)

Standalone, single-file tool that collects CI/CD variables from GitLab
groups, subgroups, and projects.  Produces a single Excel workbook with
separate sheets for each scope level plus an executive summary.

Sheets produced:
  1. Executive_Summary  - high-level counts and metadata
  2. Group_Variables    - variables from the root group only
  3. Subgroup_Variables - variables from every descendant (sub)group
  4. Project_Variables  - variables from every project in the group tree
  5. All_Variables      - combined view of all three scopes
  6. Summary            - per-scope variable statistics

Secret values are NOT exported unless --include-values is explicitly supplied.

Usage:
    python secrets_inventory.py --group my-org --token glpat-xxx
    python secrets_inventory.py --group my-org --token glpat-xxx --debug
    python secrets_inventory.py --group my-org --tokens t1,t2,t3
    python secrets_inventory.py --group my-org --token glpat-xxx --include-values

Environment variables (optional, CLI overrides take precedence):
    GITLAB_URL   - GitLab instance URL (default: https://gitlab.com)
    GITLAB_TOKEN - Personal Access Token with read_api scope
"""

import argparse
import logging
import os
import sys
import time
import threading
import requests
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION = '1.0.0'
DEFAULT_GITLAB_URL = 'https://gitlab.com'
DEFAULT_EXCEL_FILENAME = 'GitLab_Secrets_Variables_Report.xlsx'
DEFAULT_LOG_FILENAME = 'gitlab-secrets-variables-inventory.log'

# -- Field definitions per sheet -------------------------------------------

VARIABLE_FIELDS = [
    'scope_type',
    'scope_id',
    'scope_full_path',
    'parent_path',
    'depth',
    'project_id',
    'project_full_path',
    'gitlab_namespace',
    'gitlab_project',
    'key',
    'variable_type',
    'environment_scope',
    'protected',
    'masked',
    'hidden',
    'masked_and_hidden',
    'raw',
    'description',
    'value_exported',
    'variable_value',
    'scan_status',
    'scan_error',
    'scanned_at',
]

SUMMARY_FIELDS = [
    'scope_type',
    'scope_id',
    'scope_full_path',
    'parent_path',
    'depth',
    'project_id',
    'project_full_path',
    'gitlab_namespace',
    'gitlab_project',
    'variable_count',
    'masked_variable_count',
    'protected_variable_count',
    'hidden_variable_count',
    'file_variable_count',
    'env_var_count',
    'environment_scopes',
    'scan_status',
    'scan_error',
    'scanned_at',
]

EXECUTIVE_SUMMARY_FIELDS = [
    'metric',
    'value',
]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_logger = logging.getLogger('gitlab_secrets_variables_inventory')


def setup_logging(log_file=None, debug=False):
    """Configure dual logging to console and optionally to a file."""
    level = logging.DEBUG if debug else logging.INFO
    _logger.setLevel(level)

    fmt = logging.Formatter('[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(level)
    console_handler.setFormatter(fmt)
    _logger.addHandler(console_handler)

    # File handler (optional)
    if log_file:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(fmt)
        _logger.addHandler(file_handler)

    return log_file


def log(message):
    """Write an INFO-level log message."""
    _logger.info(message)


def debug_log(message, debug_enabled):
    """Write a DEBUG-level log message (shown only when debug is active)."""
    if debug_enabled:
        _logger.debug("[DEBUG] %s", message)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_int(value, default=0):
    """Convert a value to int, returning *default* on failure."""
    if value in (None, ''):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def safe_bool(value):
    """Normalize API boolean-like values."""
    if isinstance(value, bool):
        return value
    if value in (None, ''):
        return ''
    return str(value).lower() in ('1', 'true', 'yes')


def mask_token(token):
    """Mask a token for safe log output."""
    if not token or len(token) < 12:
        return '***'
    return '%s...%s' % (token[:8], token[-4:])


def join_values(values):
    """Join non-empty unique values into a semicolon-delimited string."""
    return ';'.join(sorted({str(v) for v in values if v not in (None, '')}))


def now_utc():
    """Return an ISO UTC timestamp for scan metadata."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# ---------------------------------------------------------------------------
# GitLab API Client
# ---------------------------------------------------------------------------

class GitLabApiError(Exception):
    """Raised when the GitLab API returns an error response."""

    def __init__(self, status_code, message, url=''):
        super().__init__('GitLab API error %d: %s' % (status_code, message))
        self.status_code = status_code
        self.message = message
        self.url = url


class GitLabClient:
    """GitLab REST API client with pagination, retry, and multi-token rotation."""

    def __init__(self, gitlab_url, tokens, timeout=30, retries=3):
        self.base_url = gitlab_url.rstrip('/') + '/api/v4'
        if isinstance(tokens, str):
            tokens = [tokens]
        self._tokens = list(tokens)
        self._token_index = 0
        self._token_lock = threading.Lock()
        self.timeout = timeout
        self.retries = retries

    def _next_header(self):
        """Return request headers using the next token in round-robin order."""
        with self._token_lock:
            token = self._tokens[self._token_index % len(self._tokens)]
            self._token_index += 1
        return {'PRIVATE-TOKEN': token}

    def get_json(self, path, params=None, allow_missing=False):
        """Fetch a single JSON response from the API."""
        response = self._request(path, params=params)
        if response.status_code == 404 and allow_missing:
            return None
        if response.status_code >= 400:
            raise GitLabApiError(response.status_code, response.text[:500], response.url)
        return response.json()

    def get_paginated(self, path, params=None, max_items=0, allow_missing=False):
        """Fetch all pages of a paginated API endpoint."""
        results = []
        page = 1
        base_params = dict(params or {})

        while True:
            request_params = dict(base_params)
            request_params['per_page'] = 100
            request_params['page'] = page
            response = self._request(path, params=request_params)

            if response.status_code == 404 and allow_missing:
                return results
            if response.status_code >= 400:
                raise GitLabApiError(response.status_code, response.text[:500], response.url)

            page_items = response.json()
            if not page_items:
                break

            results.extend(page_items)
            if max_items and len(results) >= max_items:
                return results[:max_items]

            total_pages = response.headers.get('X-Total-Pages')
            next_page = response.headers.get('X-Next-Page')
            if total_pages and page >= int(total_pages):
                break
            if not total_pages and not next_page and len(page_items) < 100:
                break
            page += 1

        return results

    def _request(self, path, params=None):
        """Execute a GET request with retry on transient / rate-limit errors."""
        url = self.base_url + path
        last_response = None
        for attempt in range(self.retries + 1):
            headers = self._next_header()
            try:
                response = requests.get(
                    url, headers=headers, params=params, timeout=self.timeout
                )
            except requests.exceptions.ConnectionError as exc:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Connection error on %s (attempt %d/%d) - retrying in %ds'
                    % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue
            except requests.exceptions.Timeout:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Timeout on %s (attempt %d/%d) - retrying in %ds'
                    % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue

            last_response = response
            if response.status_code not in (429, 502, 503, 504):
                return response

            if attempt >= self.retries:
                return response

            retry_after = safe_int(
                response.headers.get('Retry-After'),
                min(2 ** (attempt + 1), 60),
            )
            retry_after = max(1, retry_after)
            log('HTTP %d on %s - retrying in %ds (attempt %d/%d)'
                % (response.status_code, path, retry_after, attempt + 1, self.retries))
            time.sleep(retry_after)

        return last_response


# ---------------------------------------------------------------------------
# Discovery functions
# ---------------------------------------------------------------------------

def discover_groups(client, group_path, debug=False):
    """Discover root group and all descendant (sub)groups.

    Returns (root_group, subgroups) where root_group is the API object for
    the top-level group and subgroups is a list of API objects for every
    descendant group.
    """
    encoded = quote(group_path, safe='')
    log('Discovering groups under: %s' % group_path)

    root_group = client.get_json('/groups/%s' % encoded, allow_missing=True)
    if not root_group:
        log('ERROR: Group "%s" not found or not accessible' % group_path)
        return None, []

    descendants = client.get_paginated(
        '/groups/%s/descendant_groups' % encoded,
        params={'all_available': 'false'},
        allow_missing=True,
    )
    debug_log('Found %d descendant groups' % len(descendants), debug)

    return root_group, descendants


def discover_projects(client, group_path, include_subgroups=True,
                      include_archived=False, debug=False):
    """Discover all projects from a group (optionally including subgroups).

    Returns a list of dicts with gitlab_namespace, gitlab_project,
    gitlab_full_path, and gitlab_project_id.
    """
    encoded = quote(group_path, safe='')
    log('Discovering projects under: %s' % group_path)

    raw_projects = client.get_paginated(
        '/groups/%s/projects' % encoded,
        params={
            'include_subgroups': str(include_subgroups).lower(),
            'order_by': 'path',
            'sort': 'asc',
        },
    )

    project_rows = []
    for proj in raw_projects:
        if proj.get('archived') and not include_archived:
            continue
        full_path = (proj.get('path_with_namespace') or '').strip()
        if not full_path or '/' not in full_path:
            continue
        namespace, project_name = full_path.rsplit('/', 1)
        project_rows.append({
            'gitlab_namespace': namespace,
            'gitlab_project': project_name,
            'gitlab_full_path': full_path,
            'gitlab_project_id': proj.get('id', ''),
        })

    log('Found %d projects' % len(project_rows))
    debug_log('Project IDs: %s' % [r['gitlab_project_id'] for r in project_rows], debug)
    return project_rows


# ---------------------------------------------------------------------------
# Variable collection
# ---------------------------------------------------------------------------

def _group_depth(group, root_full_path):
    """Calculate nesting depth relative to the root group (root = 0)."""
    full_path = group.get('full_path', '')
    if full_path == root_full_path:
        return 0
    root_segments = root_full_path.count('/') + 1
    group_segments = full_path.count('/') + 1
    return group_segments - root_segments


def _group_parent(group):
    """Extract the parent path from a group's full_path."""
    full_path = group.get('full_path', '')
    if '/' in full_path:
        return full_path.rsplit('/', 1)[0]
    return ''


def _normalize_variable(variable, include_values, base_row):
    """Build one normalized variable row from an API variable object."""
    value = ''
    exported = False
    if include_values:
        value = variable.get('value', '')
        exported = True

    row = {field: '' for field in VARIABLE_FIELDS}
    row.update(base_row)
    row.update({
        'key': variable.get('key', ''),
        'variable_type': variable.get('variable_type', ''),
        'environment_scope': variable.get('environment_scope', ''),
        'protected': safe_bool(variable.get('protected')),
        'masked': safe_bool(variable.get('masked')),
        'hidden': safe_bool(variable.get('hidden')),
        'masked_and_hidden': safe_bool(variable.get('masked_and_hidden')),
        'raw': safe_bool(variable.get('raw')),
        'description': variable.get('description', ''),
        'value_exported': exported,
        'variable_value': value,
        'scan_status': 'success',
        'scan_error': '',
        'scanned_at': now_utc(),
    })
    return row


def _failed_row(scope_type, error, **extra):
    """Build a failure row when a scope cannot be scanned."""
    row = {field: '' for field in VARIABLE_FIELDS}
    row.update({
        'scope_type': scope_type,
        'scan_status': 'failed',
        'scan_error': str(error),
        'scanned_at': now_utc(),
    })
    row.update(extra)
    return row


def collect_group_variables(client, root_group, include_values, debug=False):
    """Collect CI/CD variables from the root group only.

    Returns a list of variable rows with scope_type='group'.
    """
    group_path = root_group.get('full_path', root_group.get('name', ''))
    log('Scanning root group variables: %s' % group_path)

    base = {
        'scope_type': 'group',
        'scope_id': root_group.get('id', ''),
        'scope_full_path': root_group.get('full_path', ''),
        'parent_path': '',
        'depth': 0,
    }

    rows = []
    try:
        variables = client.get_paginated(
            '/groups/%d/variables' % root_group['id'],
            allow_missing=True,
        )
        debug_log('Root group has %d variables' % len(variables), debug)
        for var in variables:
            rows.append(_normalize_variable(var, include_values, base))
    except (GitLabApiError, requests.RequestException) as err:
        log('WARNING: Failed to scan group variables for %s: %s' % (group_path, err))
        rows.append(_failed_row('group', err, scope_id=root_group.get('id', ''),
                                scope_full_path=root_group.get('full_path', '')))
    return rows


def collect_subgroup_variables(client, subgroups, root_full_path, include_values, debug=False):
    """Collect CI/CD variables from every descendant (sub)group.

    Returns a list of variable rows with scope_type='subgroup'.
    """
    rows = []
    for idx, sg in enumerate(subgroups, start=1):
        sg_path = sg.get('full_path', sg.get('name', ''))
        log('Scanning subgroup variables %d/%d: %s' % (idx, len(subgroups), sg_path))

        base = {
            'scope_type': 'subgroup',
            'scope_id': sg.get('id', ''),
            'scope_full_path': sg.get('full_path', ''),
            'parent_path': _group_parent(sg),
            'depth': _group_depth(sg, root_full_path),
        }

        try:
            variables = client.get_paginated(
                '/groups/%d/variables' % sg['id'],
                allow_missing=True,
            )
            debug_log('Subgroup %s has %d variables' % (sg_path, len(variables)), debug)
            for var in variables:
                rows.append(_normalize_variable(var, include_values, base))
        except (GitLabApiError, requests.RequestException) as err:
            log('WARNING: Failed to scan subgroup variables for %s: %s' % (sg_path, err))
            rows.append(_failed_row('subgroup', err,
                                    scope_id=sg.get('id', ''),
                                    scope_full_path=sg.get('full_path', ''),
                                    parent_path=_group_parent(sg),
                                    depth=_group_depth(sg, root_full_path)))
    return rows


def collect_project_variables(client, project_rows, include_values, debug=False):
    """Collect CI/CD variables from all projects.

    Returns a list of variable rows with scope_type='project'.
    """
    rows = []
    for idx, pr in enumerate(project_rows, start=1):
        full_path = pr.get('gitlab_full_path', '')
        log('Scanning project variables %d/%d: %s' % (idx, len(project_rows), full_path))

        base = {
            'scope_type': 'project',
            'scope_id': '',
            'scope_full_path': '',
            'parent_path': '',
            'depth': '',
            'project_id': pr.get('gitlab_project_id', ''),
            'project_full_path': full_path,
            'gitlab_namespace': pr.get('gitlab_namespace', ''),
            'gitlab_project': pr.get('gitlab_project', ''),
        }

        try:
            project = client.get_json(
                '/projects/%s' % quote(full_path, safe='')
            )
            base['project_id'] = project.get('id', pr.get('gitlab_project_id', ''))

            variables = client.get_paginated(
                '/projects/%d/variables' % project['id'],
                allow_missing=True,
            )
            debug_log('Project %s has %d variables' % (full_path, len(variables)), debug)
            for var in variables:
                rows.append(_normalize_variable(var, include_values, base))
        except (GitLabApiError, requests.RequestException) as err:
            log('WARNING: Failed to scan project variables for %s: %s' % (full_path, err))
            rows.append(_failed_row('project', err,
                                    project_id=pr.get('gitlab_project_id', ''),
                                    project_full_path=full_path,
                                    gitlab_namespace=pr.get('gitlab_namespace', ''),
                                    gitlab_project=pr.get('gitlab_project', '')))
    return rows


# ---------------------------------------------------------------------------
# Analysis / summary builders
# ---------------------------------------------------------------------------

def build_summary_rows(variable_rows):
    """Build one summary row per group/subgroup/project scope."""
    grouped = {}
    for row in variable_rows:
        st = row.get('scope_type', '')
        if st in ('group', 'subgroup'):
            key = (st, row.get('scope_id', ''), row.get('scope_full_path', ''))
        else:
            key = ('project', row.get('project_id', ''), row.get('project_full_path', ''))
        grouped.setdefault(key, []).append(row)

    summary_rows = []
    for rows in grouped.values():
        first = rows[0]
        successful = [r for r in rows if r.get('scan_status') == 'success' and r.get('key')]

        summary = {field: '' for field in SUMMARY_FIELDS}
        summary.update({
            'scope_type': first.get('scope_type', ''),
            'scope_id': first.get('scope_id', ''),
            'scope_full_path': first.get('scope_full_path', ''),
            'parent_path': first.get('parent_path', ''),
            'depth': first.get('depth', ''),
            'project_id': first.get('project_id', ''),
            'project_full_path': first.get('project_full_path', ''),
            'gitlab_namespace': first.get('gitlab_namespace', ''),
            'gitlab_project': first.get('gitlab_project', ''),
            'variable_count': len(successful),
            'masked_variable_count': sum(1 for r in successful if r.get('masked') is True),
            'protected_variable_count': sum(1 for r in successful if r.get('protected') is True),
            'hidden_variable_count': sum(
                1 for r in successful
                if r.get('hidden') is True or r.get('masked_and_hidden') is True
            ),
            'file_variable_count': sum(1 for r in successful if r.get('variable_type') == 'file'),
            'env_var_count': sum(
                1 for r in successful if r.get('variable_type') in ('env_var', '')
            ),
            'environment_scopes': join_values(
                [r.get('environment_scope') for r in successful]
            ),
            'scan_status': 'failed' if first.get('scan_status') == 'failed' else 'success',
            'scan_error': first.get('scan_error', ''),
            'scanned_at': now_utc(),
        })
        summary_rows.append(summary)

    return summary_rows


def build_executive_summary(group_rows, subgroup_rows, project_rows, summary_rows,
                            root_group, subgroups, project_source_rows, timestamp):
    """Build executive summary metrics."""
    all_rows = group_rows + subgroup_rows + project_rows
    successful = [r for r in all_rows if r.get('scan_status') == 'success' and r.get('key')]

    # Count unique environment scopes across all variables
    env_scopes = set()
    for r in successful:
        es = r.get('environment_scope', '')
        if es:
            env_scopes.add(es)

    metrics = [
        ('Report Generated', timestamp),
        ('GitLab Group', root_group.get('full_path', '') if root_group else ''),
        ('', ''),
        ('Total Groups (root)', 1 if root_group else 0),
        ('Total Subgroups', len(subgroups)),
        ('Total Projects', len(project_source_rows)),
        ('', ''),
        ('Total Variables', len(successful)),
        ('Group Variables', len([r for r in group_rows if r.get('scan_status') == 'success' and r.get('key')])),
        ('Subgroup Variables', len([r for r in subgroup_rows if r.get('scan_status') == 'success' and r.get('key')])),
        ('Project Variables', len([r for r in project_rows if r.get('scan_status') == 'success' and r.get('key')])),
        ('', ''),
        ('Masked Variables', sum(1 for r in successful if r.get('masked') is True)),
        ('Protected Variables', sum(1 for r in successful if r.get('protected') is True)),
        ('Hidden Variables', sum(
            1 for r in successful
            if r.get('hidden') is True or r.get('masked_and_hidden') is True
        )),
        ('File Variables', sum(1 for r in successful if r.get('variable_type') == 'file')),
        ('Env Var Variables', sum(
            1 for r in successful if r.get('variable_type') in ('env_var', '')
        )),
        ('Unique Environment Scopes', len(env_scopes)),
        ('Environment Scopes', join_values(env_scopes) if env_scopes else ''),
        ('', ''),
        ('Scopes with Failures', sum(
            1 for r in summary_rows if r.get('scan_status') == 'failed'
        )),
        ('Values Exported', any(r.get('value_exported') is True for r in all_rows)),
    ]

    return [{'metric': m, 'value': v} for m, v in metrics]


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel_workbook(output_file, sheet_data):
    """Write a multi-sheet Excel workbook with formatting.

    Args:
        output_file: Path for the output .xlsx file.
        sheet_data: list of (sheet_name, field_list, rows) tuples.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        log('ERROR: openpyxl is not installed. Install with: pip install openpyxl')
        return False

    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    bold_font = Font(bold=True)

    for sheet_name, fields, rows in sheet_data:
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        ws.append(fields)
        for cell in ws[1]:
            cell.font = bold_font

        # Data rows
        for row in rows:
            ws.append([row.get(f, '') for f in fields])

        # Freeze top row and add auto-filter
        ws.freeze_panes = 'A2'
        if ws.dimensions and len(rows) > 0:
            ws.auto_filter.ref = ws.dimensions

        # Auto-size columns (approximate)
        for col_idx, field in enumerate(fields, start=1):
            max_len = len(str(field))
            # Sample first 50 rows for width
            for data_row in rows[:50]:
                cell_val = str(data_row.get(field, ''))
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 60)

    wb.save(output_file)
    log('Excel workbook written: %s' % output_file)
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='GitLab CI/CD Secrets and Variables Inventory Tool - '
                    'Collects variables from groups, subgroups, and projects '
                    'into a single Excel workbook.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        '--group', required=True,
        help='GitLab group path to discover variables from',
    )
    parser.add_argument(
        '--gitlab-url',
        default=os.environ.get('GITLAB_URL', DEFAULT_GITLAB_URL),
        help='GitLab instance URL (default: GITLAB_URL env or %s)' % DEFAULT_GITLAB_URL,
    )
    parser.add_argument(
        '--token',
        default=os.environ.get('GITLAB_TOKEN', ''),
        help='GitLab Personal Access Token (default: GITLAB_TOKEN env)',
    )
    parser.add_argument(
        '--tokens',
        help='Comma-separated tokens for round-robin rate-limit rotation',
    )
    parser.add_argument(
        '--output', default=DEFAULT_EXCEL_FILENAME,
        help='Output Excel file path (default: %s)' % DEFAULT_EXCEL_FILENAME,
    )
    parser.add_argument(
        '--debug', action='store_true',
        help='Enable debug logging',
    )
    parser.add_argument(
        '--log-file',
        help='Write logs to this file in addition to console',
    )
    parser.add_argument(
        '--include-values', action='store_true',
        help='Export variable values (CAUTION: exposes secrets)',
    )
    parser.add_argument(
        '--include-archived', action='store_true',
        help='Include archived projects',
    )
    parser.add_argument(
        '--no-subgroups', action='store_true',
        help='Skip subgroup variable discovery',
    )
    parser.add_argument(
        '--skip-group-variables', action='store_true',
        help='Skip group and subgroup variable scanning',
    )
    parser.add_argument(
        '--skip-project-variables', action='store_true',
        help='Skip project variable scanning',
    )
    parser.add_argument(
        '--timeout', type=int, default=30,
        help='API request timeout in seconds (default: 30)',
    )
    parser.add_argument(
        '--retries', type=int, default=3,
        help='Retry count on transient errors (default: 3)',
    )
    parser.add_argument(
        '--version', action='version',
        version='secrets-inventory %s' % VERSION,
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Run the GitLab secrets and variables inventory collector."""
    start_time = datetime.now()
    args = parse_args()

    # Set up logging
    log_file = setup_logging(log_file=args.log_file, debug=args.debug)

    log('GitLab Secrets & Variables Inventory Tool v%s' % VERSION)

    # Resolve tokens
    tokens = []
    if args.tokens:
        tokens = [t.strip() for t in args.tokens.split(',') if t.strip()]
    if not tokens and args.token:
        tokens = [args.token]
    if not tokens:
        log('ERROR: No GitLab token provided. Set GITLAB_TOKEN env var or use --token / --tokens.')
        return 1

    log('GitLab URL: %s' % args.gitlab_url)
    log('Group:      %s' % args.group)
    log('Tokens:     %d configured' % len(tokens))
    for i, t in enumerate(tokens, start=1):
        log('  Token %d: %s' % (i, mask_token(t)))

    if args.include_values:
        log('WARNING: --include-values is enabled. Secret values will be written to output.')

    # Determine output file path
    output_file = Path(args.output)
    log('Output:     %s' % output_file)

    # Create API client
    client = GitLabClient(args.gitlab_url, tokens, timeout=args.timeout, retries=args.retries)

    # --- Phase 1: Discover groups and subgroups --------------------------
    root_group, subgroups = discover_groups(client, args.group, debug=args.debug)
    if not root_group:
        return 1

    root_full_path = root_group.get('full_path', '')

    if args.no_subgroups:
        subgroups = []
        log('Subgroup discovery skipped (--no-subgroups)')
    else:
        log('Found 1 root group, %d subgroups' % len(subgroups))

    # --- Phase 2: Discover projects --------------------------------------
    project_source_rows = discover_projects(
        client, args.group,
        include_subgroups=not args.no_subgroups,
        include_archived=args.include_archived,
        debug=args.debug,
    )

    # --- Phase 3: Collect variables --------------------------------------
    group_rows = []
    subgroup_rows = []
    project_rows = []

    if not args.skip_group_variables:
        # Root group variables
        group_rows = collect_group_variables(
            client, root_group, args.include_values, debug=args.debug
        )
        # Subgroup variables (distinct scope)
        if subgroups:
            subgroup_rows = collect_subgroup_variables(
                client, subgroups, root_full_path,
                args.include_values, debug=args.debug
            )
    else:
        log('Group/subgroup variable scanning skipped (--skip-group-variables)')

    if not args.skip_project_variables:
        project_rows = collect_project_variables(
            client, project_source_rows, args.include_values, debug=args.debug
        )
    else:
        log('Project variable scanning skipped (--skip-project-variables)')

    # --- Phase 4: Build summaries ----------------------------------------
    all_variable_rows = group_rows + subgroup_rows + project_rows
    summary_rows = build_summary_rows(all_variable_rows)

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    exec_summary = build_executive_summary(
        group_rows, subgroup_rows, project_rows, summary_rows,
        root_group, subgroups, project_source_rows, timestamp,
    )

    # --- Phase 5: Write Excel workbook -----------------------------------
    sheet_data = [
        ('Executive_Summary', EXECUTIVE_SUMMARY_FIELDS, exec_summary),
        ('Group_Variables', VARIABLE_FIELDS, group_rows),
        ('Subgroup_Variables', VARIABLE_FIELDS, subgroup_rows),
        ('Project_Variables', VARIABLE_FIELDS, project_rows),
        ('All_Variables', VARIABLE_FIELDS, all_variable_rows),
        ('Summary', SUMMARY_FIELDS, summary_rows),
    ]

    if not write_excel_workbook(output_file, sheet_data):
        return 1

    # --- Print final summary ---------------------------------------------
    elapsed = datetime.now() - start_time
    total_seconds = int(elapsed.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    successful_vars = [r for r in all_variable_rows
                       if r.get('scan_status') == 'success' and r.get('key')]

    log('=' * 60)
    log('Secrets & Variables Inventory Summary')
    log('=' * 60)
    log('  Root Group                  %s' % root_full_path)
    log('  Subgroups Scanned           %d' % len(subgroups))
    log('  Projects Scanned            %d' % len(project_source_rows))
    log('  ' + '-' * 40)
    log('  Total Variables             %d' % len(successful_vars))
    log('  Group Variables             %d' % len([r for r in group_rows if r.get('scan_status') == 'success' and r.get('key')]))
    log('  Subgroup Variables          %d' % len([r for r in subgroup_rows if r.get('scan_status') == 'success' and r.get('key')]))
    log('  Project Variables           %d' % len([r for r in project_rows if r.get('scan_status') == 'success' and r.get('key')]))
    log('  ' + '-' * 40)
    log('  Masked Variables            %d' % sum(1 for r in successful_vars if r.get('masked') is True))
    log('  Protected Variables         %d' % sum(1 for r in successful_vars if r.get('protected') is True))
    log('  Hidden Variables            %d' % sum(1 for r in successful_vars if r.get('hidden') is True or r.get('masked_and_hidden') is True))
    log('  File Variables              %d' % sum(1 for r in successful_vars if r.get('variable_type') == 'file'))
    log('  Values Exported             %s' % args.include_values)
    log('  ' + '-' * 40)
    log('  Output file:          %s' % output_file)
    if log_file:
        log('  Log file:             %s' % log_file)
    log('  Execution time:       %dh %dm %ds' % (hours, minutes, seconds))
    log('=' * 60)

    return 0


if __name__ == '__main__':
    sys.exit(main())
