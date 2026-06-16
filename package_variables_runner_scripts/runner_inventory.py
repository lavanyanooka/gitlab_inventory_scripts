#!/usr/bin/env python3
"""
GitLab Inventory Discovery Tool

Production-ready discovery tool that connects to GitLab SaaS via Personal
Access Token and performs a complete inventory discovery of all accessible
Groups, Subgroups, Projects, and Runners.  Produces a single Excel workbook
for migration assessment.

Authentication is driven by environment variables:
    GITLAB_URL   - GitLab instance URL (default: https://gitlab.com)
    GITLAB_TOKEN - Personal Access Token with read_api scope

Usage:
    # Environment variables
    export GITLAB_URL=https://gitlab.com
    export GITLAB_TOKEN=glpat-xxxxxxxxxxxxxxxxxxxx
    python inventory.py --group my-org

    # Inline override
    python inventory.py --group my-org --gitlab-url https://gitlab.com --token glpat-xxx

    # With debug logging and custom output
    python inventory.py --group my-org --debug --output report.xlsx

    # Multiple tokens for rate-limit rotation
    python inventory.py --group my-org --tokens token1,token2,token3
"""

import argparse
import logging
import os
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION = '1.0.0'
DEFAULT_GITLAB_URL = 'https://gitlab.com'
DEFAULT_OUTPUT_FILE = 'GitLab_Inventory_Report.xlsx'
DEFAULT_TIMEOUT = 30
DEFAULT_RETRIES = 3
DEFAULT_MAX_WORKERS = 4

# -- Column definitions per sheet ------------------------------------------

GROUP_COLUMNS = [
    'Group ID', 'Group Name', 'Full Path',
    'Visibility', 'Parent Group', 'Web URL',
]

SUBGROUP_COLUMNS = [
    'Subgroup ID', 'Subgroup Name', 'Parent Group',
    'Full Path', 'Visibility', 'Web URL',
]

PROJECT_COLUMNS = [
    'Project ID', 'Project Name', 'Namespace', 'Full Path',
    'Default Branch', 'Visibility', 'Archived', 'Last Activity Date',
    'Web URL',
]

RUNNER_COLUMNS = [
    'Runner ID', 'Description', 'Runner Type', 'Scope',
    'Status', 'Paused', 'Protected', 'Run Untagged', 'Locked',
    'Tags', 'Version', 'Platform', 'Architecture',
    'Maximum Timeout', 'Contacted At',
]

PROJECT_RUNNER_MAPPING_COLUMNS = [
    'Project ID', 'Project Name', 'Group Name',
    'Runner ID', 'Runner Description', 'Runner Type',
    'Runner Scope', 'Runner Tags',
]

TAG_ANALYSIS_COLUMNS = [
    'Tag Name', 'Number of Runners Using Tag', 'Number of Projects Using Tag',
]

STATISTICS_COLUMNS = [
    'Metric', 'Value',
]

EXECUTIVE_SUMMARY_COLUMNS = [
    'Metric', 'Value',
]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logger = logging.getLogger('gitlab_inventory')


def setup_logging(debug=False, log_file=None):
    """Configure logging to console and optionally to a file."""
    level = logging.DEBUG if debug else logging.INFO
    logger.setLevel(level)

    fmt = logging.Formatter(
        '[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'
    )

    # Console handler
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(level)
    console.setFormatter(fmt)
    logger.addHandler(console)

    # File handler (optional)
    if log_file:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(log_path, mode='w', encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    return logger


def log(msg):
    """Write an INFO message."""
    logger.info(msg)


def debug(msg):
    """Write a DEBUG message."""
    logger.debug('[DEBUG] %s', msg)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_int(value, default=0):
    """Convert to int, returning default on failure."""
    if value in (None, ''):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def mask_token(token):
    """Mask a token for safe log output."""
    if not token or len(token) < 12:
        return '***'
    return '%s...%s' % (token[:8], token[-4:])


# ---------------------------------------------------------------------------
# GitLab API Client
# ---------------------------------------------------------------------------

class GitLabApiError(Exception):
    """Raised when the GitLab API returns an error."""

    def __init__(self, status_code, message, url=''):
        super().__init__('GitLab API error %d: %s' % (status_code, message))
        self.status_code = status_code
        self.message = message
        self.url = url


class GitLabClient:
    """REST client with pagination, retry with exponential backoff,
    rate-limit handling, and multi-token round-robin rotation."""

    def __init__(self, gitlab_url, tokens, timeout=DEFAULT_TIMEOUT,
                 retries=DEFAULT_RETRIES):
        self.base_url = gitlab_url.rstrip('/') + '/api/v4'
        if isinstance(tokens, str):
            tokens = [tokens]
        self._tokens = list(tokens)
        self._token_index = 0
        self._token_lock = threading.Lock()
        self.timeout = timeout
        self.retries = retries

    # -- Token rotation -----------------------------------------------------

    def _next_header(self):
        """Return headers with the next token in round-robin order."""
        with self._token_lock:
            token = self._tokens[self._token_index % len(self._tokens)]
            self._token_index += 1
        return {'PRIVATE-TOKEN': token}

    # -- Public API ---------------------------------------------------------

    def get_json(self, path, params=None, allow_missing=False):
        """Fetch a single JSON object."""
        resp = self._request(path, params=params)
        if resp.status_code == 404 and allow_missing:
            return None
        if resp.status_code >= 400:
            raise GitLabApiError(resp.status_code, resp.text[:500], resp.url)
        return resp.json()

    def get_paginated(self, path, params=None, max_items=0,
                      allow_missing=False):
        """Fetch all pages from a paginated endpoint."""
        results = []
        page = 1
        base_params = dict(params or {})

        while True:
            req_params = dict(base_params)
            req_params['per_page'] = 100
            req_params['page'] = page

            resp = self._request(path, params=req_params)
            if resp.status_code == 404 and allow_missing:
                return results
            if resp.status_code >= 400:
                raise GitLabApiError(
                    resp.status_code, resp.text[:500], resp.url
                )

            items = resp.json()
            if not items:
                break

            results.extend(items)
            if max_items and len(results) >= max_items:
                return results[:max_items]

            total_pages = resp.headers.get('X-Total-Pages')
            next_page = resp.headers.get('X-Next-Page')
            if total_pages and page >= int(total_pages):
                break
            if not total_pages and not next_page and len(items) < 100:
                break
            page += 1

        return results

    # -- Internal -----------------------------------------------------------

    def _request(self, path, params=None):
        """Execute GET with retry and exponential backoff on transient
        errors (429, 502, 503, 504) and connection/timeout failures."""
        url = self.base_url + path
        last_resp = None

        for attempt in range(self.retries + 1):
            headers = self._next_header()
            try:
                resp = requests.get(
                    url, headers=headers, params=params,
                    timeout=self.timeout,
                )
            except requests.exceptions.ConnectionError as exc:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Connection error on %s (attempt %d/%d) - '
                    'retrying in %ds' % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue
            except requests.exceptions.Timeout:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Timeout on %s (attempt %d/%d) - '
                    'retrying in %ds' % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue

            last_resp = resp

            # Success or non-retryable error
            if resp.status_code not in (429, 502, 503, 504):
                return resp

            # Exhausted retries
            if attempt >= self.retries:
                return resp

            # Rate limited or transient server error - backoff
            retry_after = safe_int(
                resp.headers.get('Retry-After'),
                min(2 ** (attempt + 1), 60),
            )
            retry_after = max(1, retry_after)
            log('HTTP %d on %s - retrying in %ds (attempt %d/%d)'
                % (resp.status_code, path, retry_after,
                   attempt + 1, self.retries))
            time.sleep(retry_after)

        return last_resp


# ---------------------------------------------------------------------------
# Discovery: Groups
# ---------------------------------------------------------------------------

def discover_groups(client, group_path):
    """Discover the top-level group and all descendant groups.

    Returns (group_rows, subgroup_rows) as lists of dicts keyed by
    GROUP_COLUMNS / SUBGROUP_COLUMNS.
    """
    encoded = quote(group_path, safe='')
    log('Discovering groups under: %s' % group_path)

    top_group = client.get_json('/groups/%s' % encoded, allow_missing=True)
    if not top_group:
        log('ERROR: Group "%s" not found or not accessible' % group_path)
        return [], []

    group_rows = []
    subgroup_rows = []

    # Root group
    group_rows.append({
        'Group ID': top_group.get('id', ''),
        'Group Name': top_group.get('name', ''),
        'Full Path': top_group.get('full_path', ''),
        'Visibility': top_group.get('visibility', ''),
        'Parent Group': '',
        'Web URL': top_group.get('web_url', ''),
    })

    # Descendant groups
    descendants = client.get_paginated(
        '/groups/%s/descendant_groups' % encoded,
        params={'all_available': 'false'},
        allow_missing=True,
    )
    debug('Found %d descendant groups' % len(descendants))

    for dg in descendants:
        full_path = dg.get('full_path', '')
        parent = full_path.rsplit('/', 1)[0] if '/' in full_path else ''

        # Every descendant appears in both Groups and Subgroups sheets
        group_rows.append({
            'Group ID': dg.get('id', ''),
            'Group Name': dg.get('name', ''),
            'Full Path': full_path,
            'Visibility': dg.get('visibility', ''),
            'Parent Group': parent,
            'Web URL': dg.get('web_url', ''),
        })
        subgroup_rows.append({
            'Subgroup ID': dg.get('id', ''),
            'Subgroup Name': dg.get('name', ''),
            'Parent Group': parent,
            'Full Path': full_path,
            'Visibility': dg.get('visibility', ''),
            'Web URL': dg.get('web_url', ''),
        })

    log('Found %d groups, %d subgroups' % (len(group_rows), len(subgroup_rows)))
    return group_rows, subgroup_rows


# ---------------------------------------------------------------------------
# Discovery: Projects
# ---------------------------------------------------------------------------

def discover_projects(client, group_path):
    """Discover all projects (including subgroup projects).

    Returns (project_rows, raw_projects) where raw_projects is the list of
    API objects needed for runner mapping.
    """
    encoded = quote(group_path, safe='')
    log('Discovering projects under: %s' % group_path)

    raw_projects = client.get_paginated(
        '/groups/%s/projects' % encoded,
        params={
            'include_subgroups': 'true',
            'order_by': 'path',
            'sort': 'asc',
        },
    )

    project_rows = []
    for proj in raw_projects:
        full_path = (proj.get('path_with_namespace') or '').strip()
        namespace = full_path.rsplit('/', 1)[0] if '/' in full_path else ''
        project_rows.append({
            'Project ID': proj.get('id', ''),
            'Project Name': proj.get('name', ''),
            'Namespace': namespace,
            'Full Path': full_path,
            'Default Branch': proj.get('default_branch', ''),
            'Visibility': proj.get('visibility', ''),
            'Archived': proj.get('archived', False),
            'Last Activity Date': proj.get('last_activity_at', ''),
            'Web URL': proj.get('web_url', ''),
        })

    log('Found %d projects' % len(project_rows))
    debug('Project IDs: %s' % [p.get('id') for p in raw_projects])
    return project_rows, raw_projects


# ---------------------------------------------------------------------------
# Discovery: Runners
# ---------------------------------------------------------------------------

def _runner_summary_to_row(runner, runner_type, scope_name):
    """Convert a runner summary API object to a partial row."""
    tags = runner.get('tag_list', [])
    if not tags:
        tags = runner.get('tags', [])
    return {
        'Runner ID': runner.get('id', ''),
        'Description': runner.get('description', ''),
        'Runner Type': runner_type,
        'Scope': scope_name,
        'Status': runner.get('status', ''),
        'Paused': runner.get('paused', ''),
        'Protected': '',
        'Run Untagged': runner.get('run_untagged', ''),
        'Locked': runner.get('locked', ''),
        'Tags': '; '.join(tags) if tags else '',
        'Version': '',
        'Platform': '',
        'Architecture': '',
        'Maximum Timeout': '',
        'Contacted At': '',
    }


def _runner_detail_to_row(detail, runner_type, scope_name):
    """Convert a runner detail API object to a full row."""
    tags = detail.get('tag_list', [])
    if not tags:
        tags = detail.get('tags', [])
    return {
        'Runner ID': detail.get('id', ''),
        'Description': detail.get('description', ''),
        'Runner Type': runner_type,
        'Scope': scope_name,
        'Status': detail.get('status', ''),
        'Paused': detail.get('paused', ''),
        'Protected': detail.get('access_level', '') == 'ref_protected',
        'Run Untagged': detail.get('run_untagged', ''),
        'Locked': detail.get('locked', ''),
        'Tags': '; '.join(tags) if tags else '',
        'Version': detail.get('version', ''),
        'Platform': detail.get('platform', ''),
        'Architecture': detail.get('architecture', ''),
        'Maximum Timeout': detail.get('maximum_timeout', ''),
        'Contacted At': detail.get('contacted_at', ''),
    }


def _fetch_project_runners(client, project):
    """Fetch runners for a single project (for parallel execution)."""
    pid = project.get('id')
    pname = project.get('name', '')
    full_path = (project.get('path_with_namespace') or '').strip()
    namespace = full_path.rsplit('/', 1)[0] if '/' in full_path else ''

    runners = client.get_paginated(
        '/projects/%d/runners' % pid, allow_missing=True,
    )

    mappings = []
    runner_ids = set()
    for r in runners:
        rid = r.get('id')
        runner_ids.add(rid)
        tags = r.get('tag_list', [])
        if not tags:
            tags = r.get('tags', [])
        mappings.append({
            'Project ID': pid,
            'Project Name': pname,
            'Group Name': namespace,
            'Runner ID': rid,
            'Runner Description': r.get('description', ''),
            'Runner Type': r.get('runner_type', ''),
            'Runner Scope': '',
            'Runner Tags': '; '.join(tags) if tags else '',
        })

    return runner_ids, mappings


def discover_runners(client, group_path, raw_projects, max_workers=DEFAULT_MAX_WORKERS):
    """Discover all runners: group-level, instance-level, and per-project.

    Returns (runner_rows, mapping_rows) where runner_rows is a deduplicated
    list keyed by RUNNER_COLUMNS and mapping_rows maps projects to runners.
    """
    encoded = quote(group_path, safe='')

    # -- Group runners ------------------------------------------------------
    log('Discovering group runners...')
    group_runners = client.get_paginated(
        '/groups/%s/runners' % encoded,
        params={'type': 'group_type'},
        allow_missing=True,
    )
    debug('Found %d group runners' % len(group_runners))

    # -- Instance (shared) runners ------------------------------------------
    log('Discovering instance (shared) runners...')
    instance_runners = client.get_paginated(
        '/groups/%s/runners' % encoded,
        params={'type': 'instance_type'},
        allow_missing=True,
    )
    debug('Found %d instance runners' % len(instance_runners))

    # -- Collect unique runner IDs and summary rows -------------------------
    seen_ids = {}  # runner_id -> (runner_type, scope_name)
    for r in group_runners:
        seen_ids[r['id']] = ('group', group_path)
    for r in instance_runners:
        seen_ids[r['id']] = ('instance', 'shared')

    # -- Project runners (parallel) -----------------------------------------
    log('Discovering project runners across %d projects...' % len(raw_projects))
    all_mappings = []
    project_runner_ids = set()

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {
            pool.submit(_fetch_project_runners, client, proj): proj
            for proj in raw_projects
        }
        for future in as_completed(futures):
            try:
                rids, mappings = future.result()
                project_runner_ids.update(rids)
                all_mappings.extend(mappings)
            except Exception as exc:
                proj = futures[future]
                log('WARNING: Failed to fetch runners for project %s: %s'
                    % (proj.get('name', '?'), exc))

    # Track project-only runners
    for rid in project_runner_ids:
        if rid not in seen_ids:
            seen_ids[rid] = ('project', 'project-specific')

    # -- Fetch runner details -----------------------------------------------
    unique_ids = sorted(seen_ids.keys())
    log('Fetching details for %d unique runners...' % len(unique_ids))

    runner_rows = []
    for rid in unique_ids:
        rtype, scope = seen_ids[rid]
        try:
            detail = client.get_json('/runners/%d' % rid, allow_missing=True)
            if detail:
                runner_rows.append(_runner_detail_to_row(detail, rtype, scope))
            else:
                # Runner disappeared between list and detail fetch
                runner_rows.append({
                    'Runner ID': rid,
                    'Description': '',
                    'Runner Type': rtype,
                    'Scope': scope,
                    'Status': 'unknown',
                    'Paused': '', 'Protected': '',
                    'Run Untagged': '', 'Locked': '',
                    'Tags': '', 'Version': '', 'Platform': '',
                    'Architecture': '', 'Maximum Timeout': '',
                    'Contacted At': '',
                })
        except Exception as exc:
            log('WARNING: Could not fetch details for runner %d: %s'
                % (rid, exc))
            runner_rows.append({
                'Runner ID': rid,
                'Description': '',
                'Runner Type': rtype,
                'Scope': scope,
                'Status': 'error',
                'Paused': '', 'Protected': '',
                'Run Untagged': '', 'Locked': '',
                'Tags': '', 'Version': '', 'Platform': '',
                'Architecture': '', 'Maximum Timeout': '',
                'Contacted At': '',
            })

    log('Total unique runners: %d' % len(runner_rows))
    log('Total project-runner mappings: %d' % len(all_mappings))
    return runner_rows, all_mappings


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def build_tag_analysis(runner_rows, mapping_rows):
    """Analyze tag usage across runners and projects.

    Returns a list of dicts keyed by TAG_ANALYSIS_COLUMNS.
    """
    # Runners per tag
    tag_runner_count = {}
    for row in runner_rows:
        tags_str = row.get('Tags', '')
        if not tags_str:
            continue
        for tag in tags_str.split('; '):
            tag = tag.strip()
            if tag:
                tag_runner_count.setdefault(tag, set()).add(
                    row.get('Runner ID')
                )

    # Projects per tag (via mappings)
    tag_project_count = {}
    for row in mapping_rows:
        tags_str = row.get('Runner Tags', '')
        if not tags_str:
            continue
        for tag in tags_str.split('; '):
            tag = tag.strip()
            if tag:
                tag_project_count.setdefault(tag, set()).add(
                    row.get('Project ID')
                )

    all_tags = sorted(set(tag_runner_count) | set(tag_project_count))
    log('Unique tags discovered: %d' % len(all_tags))

    return [
        {
            'Tag Name': tag,
            'Number of Runners Using Tag': len(
                tag_runner_count.get(tag, set())
            ),
            'Number of Projects Using Tag': len(
                tag_project_count.get(tag, set())
            ),
        }
        for tag in all_tags
    ]


def build_statistics(group_rows, subgroup_rows, project_rows,
                     runner_rows, tag_analysis_rows):
    """Build summary statistics.

    Returns a list of dicts keyed by STATISTICS_COLUMNS.
    """
    instance_count = sum(
        1 for r in runner_rows if r.get('Runner Type') == 'instance'
    )
    group_count = sum(
        1 for r in runner_rows if r.get('Runner Type') == 'group'
    )
    project_count = sum(
        1 for r in runner_rows if r.get('Runner Type') == 'project'
    )
    online_count = sum(
        1 for r in runner_rows if r.get('Status') == 'online'
    )
    offline_count = sum(
        1 for r in runner_rows if r.get('Status') == 'offline'
    )

    metrics = [
        ('Total Groups', len(group_rows)),
        ('Total Subgroups', len(subgroup_rows)),
        ('Total Projects', len(project_rows)),
        ('Total Runners', len(runner_rows)),
        ('Instance Runner Count', instance_count),
        ('Group Runner Count', group_count),
        ('Project Runner Count', project_count),
        ('Online Runner Count', online_count),
        ('Offline Runner Count', offline_count),
        ('Unique Tags Count', len(tag_analysis_rows)),
    ]
    return [{'Metric': m, 'Value': v} for m, v in metrics]


def build_executive_summary(stats_rows, timestamp):
    """Build the Executive Summary from statistics plus a timestamp."""
    rows = [{'Metric': 'Report Generated', 'Value': timestamp}]
    rows.append({'Metric': '', 'Value': ''})
    rows.extend(stats_rows)
    return rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(output_path, sheet_data):
    """Write a multi-sheet Excel workbook with pandas + openpyxl formatting.

    Args:
        output_path: Path for the .xlsx file.
        sheet_data: list of (sheet_name, columns, rows) tuples.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    bold = Font(bold=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, columns, rows in sheet_data:
            # Build DataFrame with the exact column order
            df = pd.DataFrame(rows, columns=columns)
            # Fill missing columns with empty string
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            df = df[columns]

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # Bold headers
            for cell in ws[1]:
                cell.font = bold

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Auto-filter
            if ws.dimensions:
                ws.auto_filter.ref = ws.dimensions

            # Auto-size columns (sample first 50 rows)
            for col_idx, col_name in enumerate(columns, start=1):
                max_len = len(str(col_name))
                for row_idx in range(2, min(len(rows) + 2, 52)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        max_len = max(max_len, len(str(cell_val)))
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    log('Excel workbook written: %s' % output_path)
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='GitLab Inventory Discovery Tool - '
                    'Generates a complete inventory workbook for migration assessment.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        '--group', required=True,
        help='GitLab group path to inventory (e.g. my-org)',
    )
    parser.add_argument(
        '--gitlab-url',
        default=os.environ.get('GITLAB_URL', DEFAULT_GITLAB_URL),
        help='GitLab instance URL (default: GITLAB_URL env or %s)' % DEFAULT_GITLAB_URL,
    )
    parser.add_argument(
        '--token',
        default=os.environ.get('GITLAB_TOKEN', ''),
        help='GitLab Personal Access Token (default: GITLAB_TOKEN env)',
    )
    parser.add_argument(
        '--tokens',
        help='Comma-separated tokens for round-robin rate-limit rotation',
    )
    parser.add_argument(
        '--output', default=DEFAULT_OUTPUT_FILE,
        help='Output Excel file path (default: %s)' % DEFAULT_OUTPUT_FILE,
    )
    parser.add_argument(
        '--max-workers', type=int, default=DEFAULT_MAX_WORKERS,
        help='Thread pool size for parallel runner discovery (default: %d)' % DEFAULT_MAX_WORKERS,
    )
    parser.add_argument(
        '--timeout', type=int, default=DEFAULT_TIMEOUT,
        help='API request timeout in seconds (default: %d)' % DEFAULT_TIMEOUT,
    )
    parser.add_argument(
        '--retries', type=int, default=DEFAULT_RETRIES,
        help='Number of retries on transient errors (default: %d)' % DEFAULT_RETRIES,
    )
    parser.add_argument(
        '--debug', action='store_true',
        help='Enable debug logging',
    )
    parser.add_argument(
        '--log-file',
        help='Write logs to this file in addition to console',
    )
    parser.add_argument(
        '--version', action='version',
        version='gitlab-inventory %s' % VERSION,
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Entry point."""
    start_time = datetime.now()
    args = parse_args()

    # Logging
    setup_logging(debug=args.debug, log_file=args.log_file)

    log('GitLab Inventory Discovery Tool v%s' % VERSION)

    # Resolve tokens
    tokens = []
    if args.tokens:
        tokens = [t.strip() for t in args.tokens.split(',') if t.strip()]
    if not tokens and args.token:
        tokens = [args.token]
    if not tokens:
        log('ERROR: No GitLab token provided. Set GITLAB_TOKEN env var or use --token / --tokens.')
        return 1

    log('GitLab URL: %s' % args.gitlab_url)
    log('Group:      %s' % args.group)
    log('Tokens:     %d configured' % len(tokens))
    for i, t in enumerate(tokens, start=1):
        log('  Token %d: %s' % (i, mask_token(t)))
    log('Output:     %s' % args.output)

    # API client
    client = GitLabClient(
        args.gitlab_url, tokens,
        timeout=args.timeout, retries=args.retries,
    )

    # ----- Phase 1: Groups & Subgroups ------------------------------------
    group_rows, subgroup_rows = discover_groups(client, args.group)
    if not group_rows:
        log('ERROR: No groups found. Check group path and token permissions.')
        return 1

    # ----- Phase 2: Projects ----------------------------------------------
    project_rows, raw_projects = discover_projects(client, args.group)

    # ----- Phase 3: Runners -----------------------------------------------
    runner_rows, mapping_rows = discover_runners(
        client, args.group, raw_projects, max_workers=args.max_workers,
    )

    # ----- Phase 4: Analysis ----------------------------------------------
    tag_rows = build_tag_analysis(runner_rows, mapping_rows)
    stats_rows = build_statistics(
        group_rows, subgroup_rows, project_rows, runner_rows, tag_rows,
    )
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    exec_summary = build_executive_summary(stats_rows, timestamp)

    # ----- Phase 5: Write Excel -------------------------------------------
    sheet_data = [
        ('Executive_Summary', EXECUTIVE_SUMMARY_COLUMNS, exec_summary),
        ('Groups', GROUP_COLUMNS, group_rows),
        ('Subgroups', SUBGROUP_COLUMNS, subgroup_rows),
        ('Projects', PROJECT_COLUMNS, project_rows),
        ('Runners', RUNNER_COLUMNS, runner_rows),
        ('Project_Runner_Mapping', PROJECT_RUNNER_MAPPING_COLUMNS, mapping_rows),
        ('Tag_Analysis', TAG_ANALYSIS_COLUMNS, tag_rows),
        ('Statistics', STATISTICS_COLUMNS, stats_rows),
    ]

    write_excel(args.output, sheet_data)

    # ----- Summary --------------------------------------------------------
    elapsed = datetime.now() - start_time
    total_sec = int(elapsed.total_seconds())
    h, m, s = total_sec // 3600, (total_sec % 3600) // 60, total_sec % 60

    log('=' * 60)
    log('Inventory Discovery Summary')
    log('=' * 60)
    for row in stats_rows:
        log('  %-30s %s' % (row['Metric'], row['Value']))
    log('-' * 60)
    log('  Output file:          %s' % args.output)
    if args.log_file:
        log('  Log file:             %s' % args.log_file)
    log('  Execution time:       %dh %dm %ds' % (h, m, s))
    log('=' * 60)

    return 0


if __name__ == '__main__':
    sys.exit(main())
