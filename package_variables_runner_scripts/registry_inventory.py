#!/usr/bin/env python3
"""
GitLab Package Registry, Container Registry, and Artifacts Inventory Tool (Standalone)

Standalone, single-file tool that discovers all projects in a GitLab group
(including subgroups) and collects:
  1. Package Registry inventory (maven, nuget, generic, etc.)
  2. Container Registry inventory (images and tags)
  3. CI/CD Job Artifacts inventory

Outputs an Excel workbook (.xlsx) with one sheet per category.

Usage:
    python registry_inventory.py --group my-org --token glpat-xxx
    python registry_inventory.py --group my-org --token glpat-xxx --detail
    python registry_inventory.py --group my-org --token glpat-xxx --debug
    python registry_inventory.py --group my-org --tokens token1,token2,token3

Environment variables (optional, CLI overrides take precedence):
    GITLAB_URL   - GitLab instance URL (default: https://gitlab.com)
    GITLAB_TOKEN - Personal Access Token with read_api scope
"""

import argparse
import logging
import os
import sys
import time
import threading
import requests
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION = '1.0.0'
DEFAULT_GITLAB_URL = 'https://gitlab.com'
DEFAULT_EXCEL_FILENAME = 'gitlab-registry-inventory.xlsx'
DEFAULT_LOG_FILENAME = 'gitlab-registry-inventory.log'

# Package Registry detail fields (one row per version)
PACKAGE_CSV_FIELDS = [
    'package_id',
    'package_name',
    'package_type',
    'version',
    'total_versions',
    'project_id',
    'project_path',
    'group',
    'created_at',
    'last_updated',
    'size_bytes',
    'status',
]

# Package Registry summary fields (one row per package)
PACKAGE_SUMMARY_FIELDS = [
    'package_name',
    'package_type',
    'total_versions',
    'project_id',
    'project_path',
    'group',
    'earliest_created_at',
    'latest_updated',
    'total_size_bytes',
    'total_size_mb',
    'status',
]

# Container Registry fields (one row per tag)
CONTAINER_CSV_FIELDS = [
    'repository_id',
    'repository_name',
    'repository_path',
    'repository_location',
    'tag_count',
    'tag_name',
    'tag_digest',
    'tag_total_size_bytes',
    'tag_created_at',
    'project_id',
    'project_name',
    'project_path',
    'group',
    'total_size_bytes',
    'total_size_mb',
]

# Container Registry summary fields (one row per repository)
CONTAINER_SUMMARY_FIELDS = [
    'repository_id',
    'repository_name',
    'repository_path',
    'repository_location',
    'tag_count',
    'project_id',
    'project_name',
    'project_path',
    'group',
    'total_size_bytes',
    'total_size_mb',
    'earliest_tag_created_at',
    'latest_tag_created_at',
]

# Artifacts fields (one row per job with artifacts)
ARTIFACT_CSV_FIELDS = [
    'job_id',
    'job_name',
    'job_stage',
    'job_status',
    'job_ref',
    'job_ref_is_tag',
    'job_created_at',
    'job_finished_at',
    'artifact_file_name',
    'artifact_file_size_bytes',
    'artifact_expire_at',
    'artifact_count',
    'pipeline_id',
    'project_id',
    'project_name',
    'project_path',
    'group',
]

# Artifacts summary fields (one row per project)
ARTIFACT_SUMMARY_FIELDS = [
    'project_id',
    'project_name',
    'project_path',
    'group',
    'total_jobs_with_artifacts',
    'total_artifact_size_bytes',
    'total_artifact_size_mb',
    'earliest_job_created_at',
    'latest_job_created_at',
]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_logger = logging.getLogger('gitlab_registry_inventory')


def setup_logging(log_file=None, debug=False):
    """Configure dual logging to console and optionally to a file."""
    level = logging.DEBUG if debug else logging.INFO
    _logger.setLevel(level)

    fmt = logging.Formatter('[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(level)
    console_handler.setFormatter(fmt)
    _logger.addHandler(console_handler)

    # File handler (optional)
    if log_file:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(fmt)
        _logger.addHandler(file_handler)

    return log_file


def log(message):
    """Write an INFO-level log message."""
    _logger.info(message)


def debug_log(message, debug_enabled):
    """Write a DEBUG-level log message (shown only when debug is active)."""
    if debug_enabled:
        _logger.debug('[DEBUG] %s', message)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_int(value, default=0):
    """Convert a value to int, returning *default* on failure."""
    if value in (None, ''):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def mask_token(token):
    """Mask a token for safe log output."""
    if not token or len(token) < 12:
        return '***'
    return '%s...%s' % (token[:8], token[-4:])


# ---------------------------------------------------------------------------
# GitLab API Client
# ---------------------------------------------------------------------------

class GitLabApiError(Exception):
    """Raised when the GitLab API returns an error response."""

    def __init__(self, status_code, message, url=''):
        super().__init__('GitLab API error %d: %s' % (status_code, message))
        self.status_code = status_code
        self.message = message
        self.url = url


class GitLabClient:
    """GitLab REST API client with pagination, retry, and multi-token rotation."""

    def __init__(self, gitlab_url, tokens, timeout=30, retries=3):
        self.base_url = gitlab_url.rstrip('/') + '/api/v4'
        if isinstance(tokens, str):
            tokens = [tokens]
        self._tokens = list(tokens)
        self._token_index = 0
        self._token_lock = threading.Lock()
        self.timeout = timeout
        self.retries = retries

    def _next_header(self):
        """Return request headers using the next token in round-robin order."""
        with self._token_lock:
            token = self._tokens[self._token_index % len(self._tokens)]
            self._token_index += 1
        return {'PRIVATE-TOKEN': token}

    def get_json(self, path, params=None, allow_missing=False):
        """Fetch a single JSON response from the API."""
        response = self._request(path, params=params)
        if response.status_code == 404 and allow_missing:
            return None
        if response.status_code >= 400:
            raise GitLabApiError(response.status_code, response.text[:500], response.url)
        return response.json()

    def get_paginated(self, path, params=None, max_items=0, allow_missing=False):
        """Fetch all pages of a paginated API endpoint."""
        results = []
        page = 1
        base_params = dict(params or {})

        while True:
            request_params = dict(base_params)
            request_params['per_page'] = 100
            request_params['page'] = page
            response = self._request(path, params=request_params)

            if response.status_code == 404 and allow_missing:
                return results
            if response.status_code >= 400:
                raise GitLabApiError(response.status_code, response.text[:500], response.url)

            page_items = response.json()
            if not page_items:
                break

            results.extend(page_items)
            if max_items and len(results) >= max_items:
                return results[:max_items]

            total_pages = response.headers.get('X-Total-Pages')
            next_page = response.headers.get('X-Next-Page')
            if total_pages and page >= int(total_pages):
                break
            if not total_pages and not next_page and len(page_items) < 100:
                break
            page += 1

        return results

    def _request(self, path, params=None):
        """Execute a GET request with retry on transient / rate-limit errors."""
        url = self.base_url + path
        last_response = None
        for attempt in range(self.retries + 1):
            headers = self._next_header()
            try:
                response = requests.get(
                    url, headers=headers, params=params, timeout=self.timeout
                )
            except requests.exceptions.ConnectionError as exc:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Connection error on %s (attempt %d/%d) - retrying in %ds'
                    % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue
            except requests.exceptions.Timeout:
                if attempt >= self.retries:
                    raise
                wait = 2 ** attempt
                log('Timeout on %s (attempt %d/%d) - retrying in %ds'
                    % (path, attempt + 1, self.retries, wait))
                time.sleep(wait)
                continue

            last_response = response
            if response.status_code not in (429, 502, 503, 504):
                return response

            if attempt >= self.retries:
                return response

            retry_after = safe_int(
                response.headers.get('Retry-After'),
                min(2 ** (attempt + 1), 60),
            )
            retry_after = max(1, retry_after)
            log('HTTP %d on %s - retrying in %ds (attempt %d/%d)'
                % (response.status_code, path, retry_after, attempt + 1, self.retries))
            time.sleep(retry_after)

        return last_response


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

def discover_group_projects(client, group_path, debug=False):
    """Discover all projects under a group (including subgroups)."""
    encoded_group = quote(group_path, safe='')
    log('Discovering projects under group: %s' % group_path)

    projects = client.get_paginated(
        '/groups/%s/projects' % encoded_group,
        params={'include_subgroups': 'true', 'with_shared': 'false'},
    )
    log('Found %d projects in group "%s"' % (len(projects), group_path))
    debug_log('Project IDs: %s' % [p['id'] for p in projects], debug)
    return projects


def extract_top_group(project_path):
    """Extract the top-level group from a full project path."""
    parts = project_path.split('/')
    if len(parts) >= 2:
        return parts[1] if len(parts) > 2 else parts[0]
    return parts[0]


# ---------------------------------------------------------------------------
# Package Registry collection
# ---------------------------------------------------------------------------

def collect_package_versions(client, project, debug=False):
    """Fetch all package versions for a project (with file-level sizes)."""
    project_id = project['id']
    project_path = project.get('path_with_namespace', '')
    group = extract_top_group(project_path)
    rows = []

    debug_log('Fetching packages for project %d (%s)' % (project_id, project_path), debug)

    try:
        packages = client.get_paginated(
            '/projects/%d/packages' % project_id,
            allow_missing=True,
        )
    except GitLabApiError as exc:
        log('  ERROR fetching packages for %s: %s' % (project_path, exc))
        return rows

    if not packages:
        debug_log('  No packages found for %s' % project_path, debug)
        return rows

    debug_log('  Found %d packages for %s' % (len(packages), project_path), debug)

    # Group packages by name+type to compute total_versions per package
    package_groups = {}
    for pkg in packages:
        key = (pkg.get('name', ''), pkg.get('package_type', ''))
        package_groups.setdefault(key, []).append(pkg)

    # Build per-version rows
    for (pkg_name, pkg_type), versions_list in package_groups.items():
        total_versions = len(versions_list)
        for pkg in versions_list:
            # Determine package size from package_files
            size_bytes = 0
            try:
                pkg_files = client.get_json(
                    '/projects/%d/packages/%d/package_files' % (project_id, pkg['id']),
                    allow_missing=True,
                )
                if pkg_files and isinstance(pkg_files, list):
                    size_bytes = sum(safe_int(f.get('size', 0)) for f in pkg_files)
            except GitLabApiError:
                debug_log('  Could not fetch files for package %d' % pkg['id'], debug)

            row = {
                'package_id': pkg.get('id', ''),
                'package_name': pkg.get('name', ''),
                'package_type': pkg.get('package_type', ''),
                'version': pkg.get('version', ''),
                'total_versions': total_versions,
                'project_id': project_id,
                'project_path': project_path,
                'group': group,
                'created_at': pkg.get('created_at', ''),
                'last_updated': pkg.get('last_downloaded_at') or pkg.get('created_at', ''),
                'size_bytes': size_bytes,
                'status': pkg.get('status', ''),
            }
            rows.append(row)

    return rows


def collect_package_versions_light(client, project, debug=False):
    """Lightweight variant that skips file-level size lookups (size_bytes=0)."""
    project_id = project['id']
    project_path = project.get('path_with_namespace', '')
    group = extract_top_group(project_path)
    rows = []

    debug_log('Fetching packages for project %d (%s)' % (project_id, project_path), debug)

    try:
        packages = client.get_paginated(
            '/projects/%d/packages' % project_id,
            allow_missing=True,
        )
    except GitLabApiError as exc:
        log('  ERROR fetching packages for %s: %s' % (project_path, exc))
        return rows

    if not packages:
        debug_log('  No packages found for %s' % project_path, debug)
        return rows

    debug_log('  Found %d packages for %s' % (len(packages), project_path), debug)

    package_groups = {}
    for pkg in packages:
        key = (pkg.get('name', ''), pkg.get('package_type', ''))
        package_groups.setdefault(key, []).append(pkg)

    for (pkg_name, pkg_type), versions_list in package_groups.items():
        total_versions = len(versions_list)
        for pkg in versions_list:
            row = {
                'package_id': pkg.get('id', ''),
                'package_name': pkg.get('name', ''),
                'package_type': pkg.get('package_type', ''),
                'version': pkg.get('version', ''),
                'total_versions': total_versions,
                'project_id': project_id,
                'project_path': project_path,
                'group': group,
                'created_at': pkg.get('created_at', ''),
                'last_updated': pkg.get('last_downloaded_at') or pkg.get('created_at', ''),
                'size_bytes': 0,
                'status': pkg.get('status', ''),
            }
            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Container Registry collection
# ---------------------------------------------------------------------------

def collect_container_tags(client, project, debug=False):
    """Fetch all container registry repositories and tags for a project."""
    project_id = project['id']
    project_path = project.get('path_with_namespace', '')
    project_name = project.get('name', '') or project_path.rsplit('/', 1)[-1]
    group = extract_top_group(project_path)
    rows = []

    debug_log('Fetching container registries for project %d (%s)' % (project_id, project_path), debug)

    try:
        repositories = client.get_paginated(
            '/projects/%d/registry/repositories' % project_id,
            params={'tags_count': 'true', 'size': 'true'},
            allow_missing=True,
        )
    except GitLabApiError as exc:
        log('  ERROR fetching container registries for %s: %s' % (project_path, exc))
        return rows

    if not repositories:
        debug_log('  No container repositories found for %s' % project_path, debug)
        return rows

    debug_log('  Found %d container repositories for %s' % (len(repositories), project_path), debug)

    for repo in repositories:
        repo_id = repo.get('id')
        if not repo_id:
            continue

        # Derive repository_name
        repo_name = repo.get('name', '') or ''
        if not repo_name:
            repo_path = repo.get('path', '')
            repo_name = repo_path.rsplit('/', 1)[-1] if repo_path else ''

        # Fetch tags for this repository
        try:
            tags = client.get_paginated(
                '/projects/%d/registry/repositories/%d/tags' % (project_id, repo_id),
                allow_missing=True,
            )
        except GitLabApiError as exc:
            debug_log('  Could not fetch tags for repo %d: %s' % (repo_id, exc), debug)
            tags = []

        if not tags:
            # Record the repository even with no tags
            rows.append({
                'repository_id': repo_id,
                'repository_name': repo_name,
                'repository_path': repo.get('path', ''),
                'repository_location': repo.get('location', ''),
                'tag_count': 0,
                'tag_name': '',
                'tag_digest': '',
                'tag_total_size_bytes': 0,
                'tag_created_at': '',
                'project_id': project_id,
                'project_name': project_name,
                'project_path': project_path,
                'group': group,
                'total_size_bytes': 0,
                'total_size_mb': 0.0,
            })
            continue

        # Fetch tag details for size information
        tag_details = []
        for tag in tags:
            tag_name = tag.get('name', '')
            tag_detail = tag
            if tag_name:
                try:
                    detailed = client.get_json(
                        '/projects/%d/registry/repositories/%d/tags/%s'
                        % (project_id, repo_id, quote(str(tag_name), safe='')),
                        allow_missing=True,
                    )
                    if detailed:
                        tag_detail = detailed
                except GitLabApiError:
                    debug_log('  Could not fetch detail for tag %s' % tag_name, debug)
            tag_details.append(tag_detail)

        # Compute total repo size from tag sizes
        repo_total_size = sum(safe_int(td.get('total_size')) for td in tag_details)

        for tag_detail in tag_details:
            tag_size = safe_int(tag_detail.get('total_size'))
            rows.append({
                'repository_id': repo_id,
                'repository_name': repo_name,
                'repository_path': repo.get('path', ''),
                'repository_location': repo.get('location', ''),
                'tag_count': safe_int(repo.get('tags_count', len(tags))),
                'tag_name': tag_detail.get('name', ''),
                'tag_digest': tag_detail.get('digest', ''),
                'tag_total_size_bytes': tag_size,
                'tag_created_at': tag_detail.get('created_at', ''),
                'project_id': project_id,
                'project_name': project_name,
                'project_path': project_path,
                'group': group,
                'total_size_bytes': repo_total_size,
                'total_size_mb': round(repo_total_size / (1024 * 1024), 2),
            })

    return rows


# ---------------------------------------------------------------------------
# Artifacts collection
# ---------------------------------------------------------------------------

def collect_artifacts(client, project, debug=False):
    """Fetch all CI/CD jobs that have artifacts for a project."""
    project_id = project['id']
    project_path = project.get('path_with_namespace', '')
    project_name = project.get('name', '') or project_path.rsplit('/', 1)[-1]
    group = extract_top_group(project_path)
    rows = []

    debug_log('Fetching jobs with artifacts for project %d (%s)' % (project_id, project_path), debug)

    try:
        jobs = client.get_paginated(
            '/projects/%d/jobs' % project_id,
            params={'order_by': 'id', 'sort': 'desc'},
            allow_missing=True,
        )
    except GitLabApiError as exc:
        log('  ERROR fetching jobs for %s: %s' % (project_path, exc))
        return rows

    if not jobs:
        debug_log('  No jobs found for %s' % project_path, debug)
        return rows

    debug_log('  Found %d total jobs for %s' % (len(jobs), project_path), debug)

    for job in jobs:
        # Only include jobs that have artifacts
        artifacts_file = job.get('artifacts_file') or {}
        artifacts_list = job.get('artifacts') or []
        artifact_file_size = safe_int(artifacts_file.get('size'))
        has_artifacts = (
            bool(artifacts_list)
            or artifact_file_size > 0
            or bool(artifacts_file.get('filename'))
        )
        if not has_artifacts:
            continue

        pipeline = job.get('pipeline') or {}
        row = {
            'job_id': job.get('id', ''),
            'job_name': job.get('name', ''),
            'job_stage': job.get('stage', ''),
            'job_status': job.get('status', ''),
            'job_ref': job.get('ref', ''),
            'job_ref_is_tag': job.get('tag', ''),
            'job_created_at': job.get('created_at', ''),
            'job_finished_at': job.get('finished_at', ''),
            'artifact_file_name': artifacts_file.get('filename', ''),
            'artifact_file_size_bytes': artifact_file_size,
            'artifact_expire_at': job.get('artifacts_expire_at', ''),
            'artifact_count': len(artifacts_list),
            'pipeline_id': pipeline.get('id', ''),
            'project_id': project_id,
            'project_name': project_name,
            'project_path': project_path,
            'group': group,
        }
        rows.append(row)

    if rows:
        debug_log('  Found %d jobs with artifacts for %s' % (len(rows), project_path), debug)

    return rows


# ---------------------------------------------------------------------------
# Aggregation / Summary
# ---------------------------------------------------------------------------

def aggregate_packages(detail_rows):
    """Collapse per-version package rows into one row per unique package."""
    groups = {}
    for row in detail_rows:
        key = (row['package_name'], row['package_type'], row['project_id'])
        groups.setdefault(key, []).append(row)

    summary_rows = []
    for (pkg_name, pkg_type, proj_id), versions in groups.items():
        total_size = sum(safe_int(v['size_bytes']) for v in versions)
        created_dates = [v['created_at'] for v in versions if v.get('created_at')]
        updated_dates = [v['last_updated'] for v in versions if v.get('last_updated')]

        summary_rows.append({
            'package_name': pkg_name,
            'package_type': pkg_type,
            'total_versions': len(versions),
            'project_id': proj_id,
            'project_path': versions[0].get('project_path', ''),
            'group': versions[0].get('group', ''),
            'earliest_created_at': min(created_dates) if created_dates else '',
            'latest_updated': max(updated_dates) if updated_dates else '',
            'total_size_bytes': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'status': versions[0].get('status', ''),
        })

    return summary_rows


def aggregate_containers(detail_rows):
    """Collapse per-tag container rows into one row per repository."""
    groups = {}
    for row in detail_rows:
        key = (row['repository_id'], row['project_id'])
        groups.setdefault(key, []).append(row)

    summary_rows = []
    for (repo_id, proj_id), tags in groups.items():
        tag_dates = [t['tag_created_at'] for t in tags if t.get('tag_created_at')]
        total_size = safe_int(tags[0].get('total_size_bytes'))

        summary_rows.append({
            'repository_id': repo_id,
            'repository_name': tags[0].get('repository_name', ''),
            'repository_path': tags[0].get('repository_path', ''),
            'repository_location': tags[0].get('repository_location', ''),
            'tag_count': len([t for t in tags if t.get('tag_name')]),
            'project_id': proj_id,
            'project_name': tags[0].get('project_name', ''),
            'project_path': tags[0].get('project_path', ''),
            'group': tags[0].get('group', ''),
            'total_size_bytes': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'earliest_tag_created_at': min(tag_dates) if tag_dates else '',
            'latest_tag_created_at': max(tag_dates) if tag_dates else '',
        })

    return summary_rows


def aggregate_artifacts(detail_rows):
    """Collapse per-job artifact rows into one row per project."""
    groups = {}
    for row in detail_rows:
        groups.setdefault(row['project_id'], []).append(row)

    summary_rows = []
    for proj_id, jobs in groups.items():
        total_size = sum(safe_int(j['artifact_file_size_bytes']) for j in jobs)
        created_dates = [j['job_created_at'] for j in jobs if j.get('job_created_at')]

        summary_rows.append({
            'project_id': proj_id,
            'project_name': jobs[0].get('project_name', ''),
            'project_path': jobs[0].get('project_path', ''),
            'group': jobs[0].get('group', ''),
            'total_jobs_with_artifacts': len(jobs),
            'total_artifact_size_bytes': total_size,
            'total_artifact_size_mb': round(total_size / (1024 * 1024), 2),
            'earliest_job_created_at': min(created_dates) if created_dates else '',
            'latest_job_created_at': max(created_dates) if created_dates else '',
        })

    return summary_rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel_workbook(output_file, sheet_data):
    """Write an Excel workbook with multiple sheets.

    *sheet_data* is a list of (sheet_name, fieldnames, rows) tuples.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        log('ERROR: openpyxl is not installed. Install with: pip install openpyxl')
        return False

    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for sheet_name, fieldnames, rows in sheet_data:
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row in rows:
            ws.append([row.get(field, '') for field in fieldnames])

        # Freeze header row and enable auto-filter
        ws.freeze_panes = 'A2'
        if ws.dimensions:
            ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='GitLab Package Registry, Container Registry, and Artifacts Inventory Tool - '
                    'Collects registry and artifact data into a single Excel workbook.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        '--group', required=True,
        help='GitLab group path to scan',
    )
    parser.add_argument(
        '--gitlab-url',
        default=os.environ.get('GITLAB_URL', DEFAULT_GITLAB_URL),
        help='GitLab instance URL (default: GITLAB_URL env or %s)' % DEFAULT_GITLAB_URL,
    )
    parser.add_argument(
        '--token',
        default=os.environ.get('GITLAB_TOKEN', ''),
        help='GitLab Personal Access Token (default: GITLAB_TOKEN env)',
    )
    parser.add_argument(
        '--tokens',
        help='Comma-separated tokens for round-robin rate-limit rotation',
    )
    parser.add_argument(
        '--output', default=DEFAULT_EXCEL_FILENAME,
        help='Output Excel file path (default: %s)' % DEFAULT_EXCEL_FILENAME,
    )
    parser.add_argument(
        '--debug', action='store_true',
        help='Enable debug logging',
    )
    parser.add_argument(
        '--log-file',
        help='Write logs to this file in addition to console',
    )
    parser.add_argument(
        '--skip-files', action='store_true',
        help='Skip fetching individual package files (faster, but size_bytes will be 0)',
    )
    parser.add_argument(
        '--detail', action='store_true',
        help='Show per-version/per-tag/per-job detail rows instead of aggregated summary',
    )
    parser.add_argument(
        '--skip-packages', action='store_true',
        help='Skip package registry scanning',
    )
    parser.add_argument(
        '--skip-containers', action='store_true',
        help='Skip container registry scanning',
    )
    parser.add_argument(
        '--skip-artifacts', action='store_true',
        help='Skip CI/CD artifacts scanning',
    )
    parser.add_argument(
        '--timeout', type=int, default=30,
        help='API request timeout in seconds (default: 30)',
    )
    parser.add_argument(
        '--retries', type=int, default=3,
        help='Retry count on transient errors (default: 3)',
    )
    parser.add_argument(
        '--version', action='version',
        version='registry-inventory %s' % VERSION,
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Entry point."""
    script_start_time = datetime.now()
    args = parse_args()

    # Set up logging
    log_file = setup_logging(log_file=args.log_file, debug=args.debug)

    log('GitLab Registry Inventory Tool v%s' % VERSION)

    # Resolve tokens
    tokens = []
    if args.tokens:
        tokens = [t.strip() for t in args.tokens.split(',') if t.strip()]
    if not tokens and args.token:
        tokens = [args.token]
    if not tokens:
        log('ERROR: No GitLab token provided. Set GITLAB_TOKEN env var or use --token / --tokens.')
        return 1

    debug = args.debug
    skip_files = args.skip_files
    detail_mode = args.detail
    skip_packages = args.skip_packages
    skip_containers = args.skip_containers
    skip_artifacts = args.skip_artifacts

    log('GitLab URL: %s' % args.gitlab_url)
    log('Group:      %s' % args.group)
    log('Tokens:     %d configured' % len(tokens))
    for i, t in enumerate(tokens, start=1):
        log('  Token %d: %s' % (i, mask_token(t)))

    # Log which scans are enabled
    scans = []
    if not skip_packages:
        scans.append('Package Registry')
    if not skip_containers:
        scans.append('Container Registry')
    if not skip_artifacts:
        scans.append('Artifacts')
    log('Scans:      %s' % ', '.join(scans))
    if skip_files:
        log('Package file fetching SKIPPED (--skip-files). size_bytes will be 0.')
    if detail_mode:
        log('Detail mode: sheets will show per-version/per-tag/per-job rows.')
    else:
        log('Summary mode (default): one row per package/repo/project.')

    output_file = Path(args.output)
    log('Output:     %s' % output_file)

    # API client
    client = GitLabClient(args.gitlab_url, tokens, timeout=args.timeout, retries=args.retries)

    # Discover projects
    projects = discover_group_projects(client, args.group, debug=debug)
    if not projects:
        log('No projects found. Exiting.')
        return 0

    # Collect inventories per project
    all_pkg_rows = []
    all_container_rows = []
    all_artifact_rows = []
    failed_projects = []

    for idx, project in enumerate(projects):
        project_path = project.get('path_with_namespace', 'id:%d' % project['id'])
        log('[%d/%d] Scanning: %s' % (idx + 1, len(projects), project_path))

        try:
            # Package Registry
            if not skip_packages:
                if skip_files:
                    pkg_rows = collect_package_versions_light(client, project, debug=debug)
                else:
                    pkg_rows = collect_package_versions(client, project, debug=debug)
                if pkg_rows:
                    log('  Packages: %d version(s)' % len(pkg_rows))
                all_pkg_rows.extend(pkg_rows)

            # Container Registry
            if not skip_containers:
                container_rows = collect_container_tags(client, project, debug=debug)
                if container_rows:
                    log('  Containers: %d tag(s)' % len(container_rows))
                all_container_rows.extend(container_rows)

            # Artifacts
            if not skip_artifacts:
                artifact_rows = collect_artifacts(client, project, debug=debug)
                if artifact_rows:
                    log('  Artifacts: %d job(s)' % len(artifact_rows))
                all_artifact_rows.extend(artifact_rows)

        except GitLabApiError as exc:
            log('  ERROR: %s' % exc)
            failed_projects.append(project_path)
        except requests.exceptions.RequestException as exc:
            log('  ERROR (network): %s' % exc)
            failed_projects.append(project_path)

        if (idx + 1) % 10 == 0:
            log('  Progress: %d/%d projects scanned' % (idx + 1, len(projects)))

    # Build output data
    # Default is summary (aggregated) mode; --detail gives per-version rows
    if detail_mode:
        pkg_fields = PACKAGE_CSV_FIELDS
        pkg_out = all_pkg_rows
        container_fields = CONTAINER_CSV_FIELDS
        container_out = all_container_rows
        artifact_fields = ARTIFACT_CSV_FIELDS
        artifact_out = all_artifact_rows
    else:
        pkg_fields = PACKAGE_SUMMARY_FIELDS
        pkg_out = aggregate_packages(all_pkg_rows)
        container_fields = CONTAINER_SUMMARY_FIELDS
        container_out = aggregate_containers(all_container_rows)
        artifact_fields = ARTIFACT_SUMMARY_FIELDS
        artifact_out = aggregate_artifacts(all_artifact_rows)
        log('Aggregated: %d pkg versions -> %d packages, '
            '%d tags -> %d repos, '
            '%d jobs -> %d projects'
            % (len(all_pkg_rows), len(pkg_out),
               len(all_container_rows), len(container_out),
               len(all_artifact_rows), len(artifact_out)))

    # Write Excel workbook
    sheet_data = []
    if not skip_packages:
        sheet_name = 'Package Registry (Detail)' if detail_mode else 'Package Registry'
        sheet_data.append((sheet_name, pkg_fields, pkg_out))
    if not skip_containers:
        sheet_name = 'Container Registry (Detail)' if detail_mode else 'Container Registry'
        sheet_data.append((sheet_name, container_fields, container_out))
    if not skip_artifacts:
        sheet_name = 'Artifacts (Detail)' if detail_mode else 'Artifacts'
        sheet_data.append((sheet_name, artifact_fields, artifact_out))

    try:
        if write_excel_workbook(output_file, sheet_data):
            log('Excel workbook written: %s' % output_file)
        else:
            log('ERROR: Failed to write Excel workbook (openpyxl missing?)')
            return 1
    except PermissionError:
        log('ERROR: Cannot write to %s - file may be open. Close it and retry.' % output_file)
        return 1
    except Exception as exc:
        log('ERROR: Failed to write Excel workbook: %s' % exc)
        return 1

    # Summary report
    script_end_time = datetime.now()
    execution_time = script_end_time - script_start_time
    total_seconds = int(execution_time.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    log('=' * 60)
    log('Summary')
    log('=' * 60)
    log('  Projects scanned:     %d' % len(projects))
    log('  Failed projects:      %d' % len(failed_projects))
    if failed_projects:
        for fp in failed_projects:
            log('    - %s' % fp)

    if not skip_packages:
        unique_pkgs = set()
        pkg_type_counts = {}
        pkg_total_size = 0
        for row in all_pkg_rows:
            unique_pkgs.add((row['project_id'], row['package_name'], row['package_type']))
            ptype = row['package_type']
            pkg_type_counts[ptype] = pkg_type_counts.get(ptype, 0) + 1
            pkg_total_size += safe_int(row['size_bytes'])
        log('  Package versions:     %d' % len(all_pkg_rows))
        log('  Unique packages:      %d' % len(unique_pkgs))
        log('  Package total size:   %.2f MB' % (pkg_total_size / (1024 * 1024)))
        if pkg_type_counts:
            for ptype, count in sorted(pkg_type_counts.items()):
                log('    %s: %d version(s)' % (ptype, count))

    if not skip_containers:
        unique_repos = set()
        container_total_size = 0
        total_tags = 0
        for row in all_container_rows:
            unique_repos.add(row['repository_id'])
            container_total_size += safe_int(row.get('tag_total_size_bytes'))
            if row.get('tag_name'):
                total_tags += 1
        log('  Container repos:      %d' % len(unique_repos))
        log('  Container tags:       %d' % total_tags)
        log('  Container total size: %.2f MB' % (container_total_size / (1024 * 1024)))

    if not skip_artifacts:
        artifact_total_size = sum(safe_int(r['artifact_file_size_bytes']) for r in all_artifact_rows)
        log('  Jobs with artifacts:  %d' % len(all_artifact_rows))
        log('  Artifact total size:  %.2f MB' % (artifact_total_size / (1024 * 1024)))

    log('  Output file:          %s' % output_file)
    if log_file:
        log('  Log file:             %s' % log_file)
    log('  Execution time:       %dh %dm %ds' % (hours, minutes, seconds))
    log('=' * 60)

    if failed_projects:
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
